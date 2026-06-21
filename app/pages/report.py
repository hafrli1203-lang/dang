"""Screen 3 -- 성과 입력 + 보고서 생성."""
import asyncio
import io
from pathlib import Path
from typing import List, Dict

from nicegui import ui, app as nicegui_app

from app.common import create_nav, next_step_bar
from app.theme import section_header
from app.export_manager import ExportManager
from app.paths import CHARTS_DIR, sanitize_filename
from app.database import (
    get_project,
    get_projects,
    get_performance_rows,
    save_performance_rows,
    get_latest_report,
    save_report_content,
)
from app.ai_engine import build_report_prompt, calc_kpi, SYSTEM_GUIDE_REPORT
from app import store_wiki
from app import knowledge
from app.ai.providers import get_provider, ClaudeProvider, OpenAIProvider
from app.reporting.docx_report import build_report_docx
from app.reporting.parsers import parse_daangn_csv
from app.reporting.demographic import parse_demographic_xlsx
from app.funnel_widget import build_funnel_stages, render_funnel, FULL_METRICS
from app.chart_preview import make_charts


# -- helpers --

def _parse_int(val) -> int:
    try:
        return int(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0


_EXPECTED_COLS = ("기간", "비용", "노출", "클릭", "문의", "단골", "쿠폰")


def _validate_excel_header(header_row: tuple) -> str | None:
    if not header_row or len(header_row) < 7:
        return f"열이 7개보다 적어요 (지금 {len(header_row) if header_row else 0}개). 기간|비용|노출|클릭|문의|단골|쿠폰 순서로 맞춰 주세요."
    h = [str(c or "").strip().replace("(원)", "").replace("(명)", "").replace("(건)", "").replace("(회)", "") for c in header_row[:7]]
    for idx, expected in enumerate(_EXPECTED_COLS):
        if expected not in h[idx]:
            return f"열 {idx+1} 이름이 '{h[idx]}'예요. '{expected}'이(가) 들어간 이름으로 바꿔 주세요."
    return None


def _parse_excel(content: bytes) -> tuple[List[Dict], str | None]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_out = []
    warning = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            warning = _validate_excel_header(row)
            continue
        if not any(row):
            continue
        rows_out.append({
            "period_label": str(row[0]) if row[0] is not None else f"기간{i}",
            "cost": _parse_int(row[1]),
            "impressions": _parse_int(row[2]),
            "clicks": _parse_int(row[3]),
            "inquiries": _parse_int(row[4]),
            "regulars": _parse_int(row[5] if len(row) > 5 else 0),
            "coupons": _parse_int(row[6] if len(row) > 6 else 0),
        })
    return rows_out, warning


def _rows_from_daangn_breakdown(file_bytes: bytes) -> tuple[List[Dict], dict] | None:
    """당근 광고관리자 직접 내보내기(long-format) xlsx면 날짜별로 롤업해서 rows 반환.

    7열 템플릿(기간|비용|노출|클릭|문의|단골|쿠폰)이면 timeseries가 비므로 None을
    돌려 _parse_excel 폴백을 타게 한다. 파일에 없는 지표(문의/단골/쿠폰 등)는 0으로
    두되, metrics_available로 '실제 측정된 지표'를 함께 돌려준다(있는 것만 솔직히).
    """
    try:
        parsed = parse_demographic_xlsx(file_bytes)
    except Exception as exc:  # noqa: BLE001
        _report_log.debug("long-format breakdown 파싱 실패, 기본 파서로 폴백: %s", exc)
        return None
    ts = parsed.get("timeseries") or []
    if not ts:
        return None
    rows = [{
        "period_label": r["date"],
        "cost": r["cost"],
        "impressions": r["impressions"],
        "clicks": r["clicks"],
        "inquiries": r["inquiries"],
        "regulars": r["regulars"],
        "coupons": r["coupons"],
    } for r in ts]
    # 세그먼트 심화(캠페인 판정) 요약 — 슬라이드 보고서의 '세그먼트' 장에 쓴다.
    segment_rows: List[Dict] = []
    try:
        from app.reporting.demographic import judge_campaigns
        campaigns = parsed.get("campaigns") or []
        judged = judge_campaigns(campaigns)
        judged.sort(key=lambda j: j.campaign.cost, reverse=True)
        for j in judged[:8]:
            segment_rows.append({
                "label": j.campaign.name,
                "cost": j.campaign.cost,
                "cpa": j.campaign.cpa,
                "verdict": j.verdict,
            })
    except Exception:  # noqa: BLE001
        segment_rows = []
    info = {
        "metrics_available": parsed.get("metrics_available", []),
        "campaign_names": parsed.get("meta", {}).get("campaign_names", []),
        "segment_rows": segment_rows,
    }
    return rows, info


def _blank_row(idx: int) -> Dict:
    return {
        "period_label": f"기간{idx}",
        "cost": 0, "impressions": 0, "clicks": 0,
        "inquiries": 0, "regulars": 0, "coupons": 0,
    }


# -- docx_report bridge helpers --

def _rows_to_timeseries(rows: List[Dict]) -> List[Dict]:
    return [{
        "date": r.get("period_label", ""),
        "spend": r.get("cost", 0),
        "clicks": r.get("clicks", 0),
        "chats": r.get("inquiries", 0),
        "impressions": r.get("impressions", 0),
        "followers": r.get("regulars", 0),
        "coupons": r.get("coupons", 0),
    } for r in rows]


def _kpi_to_new(kpi: dict) -> dict:
    return {
        "total_spend": kpi.get("total_cost", 0),
        "total_impressions": kpi.get("total_impressions", 0),
        "total_clicks": kpi.get("total_clicks", 0),
        "total_chats": kpi.get("total_inquiries", 0),
        "total_followers": kpi.get("total_regulars", 0),
        "total_coupons": kpi.get("total_coupons", 0),
        "ctr": kpi.get("ctr", 0.0),
        "cpc": kpi.get("cpc", 0.0),
        "cpa": kpi.get("cpa", 0.0),
        "cpm": kpi.get("cpm", 0.0),
        "cpr": kpi.get("cpr", 0.0),
        "cp_coupon": kpi.get("cp_coupon", 0.0),
        "cvr_click_inquiry": kpi.get("cvr_click_inquiry", 0.0),
        "cvr_click_regular": kpi.get("cvr_click_regular", 0.0),
        "cvr_inquiry_regular": kpi.get("cvr_inquiry_regular", 0.0),
        "cvr_inquiry_coupon": kpi.get("cvr_inquiry_coupon", 0.0),
        "cvr_regular_coupon": kpi.get("cvr_regular_coupon", 0.0),
    }


def _extract_list_items(text: str) -> List[str]:
    import re
    items: List[str] = []
    current: List[str] = []
    for line in text.split("\n"):
        m = re.match(r"^\s*(?:\d+[.)]\s+|\*\s+|-\s+|•\s+)(.+)", line)
        if m:
            if current:
                items.append(" ".join(current))
            current = [m.group(1).strip()]
        elif line.strip() and current:
            current.append(line.strip())
        elif not line.strip() and current:
            items.append(" ".join(current))
            current = []
    if current:
        items.append(" ".join(current))
    return [i for i in items if i]


def _parse_ai_insights(content: str) -> dict:
    import re
    import json

    result = {
        "conclusion": "", "next_actions": [], "good": "", "blocked": "",
        "hypothesis": "", "experiments": [], "judgment": {},
        "summary": "", "insights": [], "actions": [],
    }

    json_match = re.search(r"```json\s*\n(.*?)\n\s*```", content, re.DOTALL)
    if json_match:
        try:
            parsed = json.loads(json_match.group(1))
            if isinstance(parsed, dict):
                for key in result:
                    if key in parsed:
                        result[key] = parsed[key]
                if isinstance(result.get("judgment"), dict):
                    kr_map = {"확대": "expand", "검토": "review", "중단": "stop"}
                    normalized = {}
                    for k, v in result["judgment"].items():
                        normalized[kr_map.get(k, k)] = v
                    result["judgment"] = normalized
                if not result["summary"]:
                    result["summary"] = result["conclusion"]
                if not result["actions"] and result["next_actions"]:
                    result["actions"] = result["next_actions"]
                if not result["next_actions"] and result["actions"]:
                    result["next_actions"] = result["actions"]
                return result
        except (json.JSONDecodeError, TypeError):
            pass

    parts = re.split(r"(?m)^## ", content)

    for part in parts:
        if not part.strip():
            continue
        first_line, _, body = part.partition("\n")
        header = first_line.strip().lower()
        body = body.strip()

        if "결론" in header:
            result["conclusion"] = body
        elif "next action" in header or "다음 액션" in header:
            items = _extract_list_items(body)
            result["next_actions"] = items if items else ([body] if body else [])
        elif "잘 된" in header:
            result["good"] = body
        elif "막힌" in header:
            result["blocked"] = body
        elif "가설" in header:
            result["hypothesis"] = body
        elif "실험" in header:
            experiments = []
            for line in body.split("\n"):
                line = line.strip()
                if not line:
                    continue
                line = re.sub(r"^\s*\d+[.)]\s*", "", line)
                if "|" in line:
                    parts_pipe = [p.strip() for p in line.split("|")]
                    if len(parts_pipe) >= 5:
                        experiments.append({
                            "priority": parts_pipe[0], "change": parts_pipe[1],
                            "success_criteria": parts_pipe[2], "owner": parts_pipe[3],
                            "schedule": parts_pipe[4],
                        })
                    else:
                        experiments.append({
                            "priority": "-", "change": line,
                            "success_criteria": "-", "owner": "-", "schedule": "-",
                        })
                elif line:
                    experiments.append({
                        "priority": "-", "change": line,
                        "success_criteria": "-", "owner": "-", "schedule": "-",
                    })
            result["experiments"] = experiments
        elif "판단" in header:
            judgment = {}
            key_map = {"확대": "expand", "검토": "review", "중단": "stop"}
            for line in body.split("\n"):
                line = line.strip()
                m = re.match(
                    r"[-•]?\s*\*{0,2}(확대|검토|중단)\*{0,2}\s*[:：]\s*(.+)", line,
                )
                if m:
                    judgment[key_map[m.group(1)]] = m.group(2).strip()
            result["judgment"] = judgment
        elif "요약" in header and "인사이트" not in header:
            result["summary"] = body
            if not result["conclusion"]:
                result["conclusion"] = body
        elif "인사이트" in header:
            items = _extract_list_items(body)
            result["insights"] = items if items else ([body] if body else [])
        elif "액션" in header or "action" in header:
            items = _extract_list_items(body)
            result["actions"] = items if items else ([body] if body else [])
            if not result["next_actions"]:
                result["next_actions"] = result["actions"]

    if not result["conclusion"]:
        result["conclusion"] = result.get("summary") or content[:600]
    if not result["summary"]:
        result["summary"] = result["conclusion"]
    if not result["actions"] and result["next_actions"]:
        result["actions"] = result["next_actions"]

    return result


_report_log = __import__("logging").getLogger("report")


def _make_report_docx_bytes(project: dict, rows: List[Dict], kpi: dict, content: str) -> bytes:
    import tempfile
    import time as _time
    _report_log.info("DOCX 생성 시작 (rows=%d)", len(rows))
    t0 = _time.monotonic()
    with tempfile.TemporaryDirectory() as tmp_dir:
        out = Path(tmp_dir) / "report.docx"
        meta_keys = (
            "name", "period", "goal", "industry", "region", "budget",
            "campaign_name", "author", "target", "operation_method", "benefits",
        )
        _report_log.info("차트 + DOCX 빌드 시작...")
        build_report_docx(
            project_meta={k: project.get(k, "") for k in meta_keys},
            kpi=_kpi_to_new(kpi),
            timeseries=_rows_to_timeseries(rows),
            insights=_parse_ai_insights(content),
            output_path=out,
            chart_dir=CHARTS_DIR,
        )
        elapsed = _time.monotonic() - t0
        _report_log.info("DOCX 생성 완료 (%.1f초)", elapsed)
        return out.read_bytes()


def _create_sample_excel() -> None:
    try:
        import openpyxl
    except ImportError:
        ui.notify("엑셀 템플릿 기능에 openpyxl 설치가 필요해요. 터미널에서 pip install openpyxl을 실행해 주세요.", type="negative")
        return
    try:
        from app.paths import EXPORTS_DIR
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "성과데이터"
        headers = ["기간", "비용(원)", "노출", "클릭", "문의", "단골", "쿠폰"]
        ws.append(headers)
        sample_rows = [
            ["1주차", 75000, 12000, 480, 18, 3, 5],
            ["2주차", 75000, 13500, 540, 21, 4, 7],
            ["3주차", 75000, 11800, 420, 16, 2, 4],
            ["4주차", 75000, 14200, 610, 25, 5, 9],
        ]
        for r in sample_rows:
            ws.append(r)
        downloads = Path.home() / "Downloads"
        out_dir = downloads if downloads.exists() else EXPORTS_DIR
        out = out_dir / "당근광고_성과템플릿.xlsx"
        wb.save(out)
        ui.notify(f"템플릿을 저장했어요: {out}", type="positive")
    except Exception as exc:
        ui.notify(f"템플릿을 만들지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative")


# -- Page --

def render_report_body() -> None:
    """성과 보고서 본문(create_nav 제외). /report 병합 페이지의 '성과 요약' 탭에서 재사용."""
    page_state: dict = {
        "rows": [], "kpi": {}, "report_content": "",
        "engine": "claude", "c_text": "", "g_text": "", "cancelled": False,
        "metrics_available": set(FULL_METRICS),
        "segment_rows": [],
    }

    with ui.column().classes("dg-page-content w-full gap-5"):
        next_step_bar("/report")  # CSS order로 본문 맨 아래에 '다음 단계' 흐름 버튼

        # Page header
        ui.label("성과 보고서").classes("dg-page-title")
        ui.label("광고 성과 데이터를 분석하고 AI 보고서를 만들어 드려요.").classes("dg-page-subtitle")

        # 프로젝트 선택은 사이드바 컨텍스트 스위처(매장→캠페인)로 일원화됨.
        # 페이지 로드 시 current_project_id 기준으로 _load_saved_data가 자동 실행된다.

        # -- Data Input --
        with ui.card().classes("dg-card w-full"):
            section_header("upload_file", "성과 데이터 입력", "CSV/XLSX 파일을 올리거나 직접 입력해 주세요.")

            # 입력 평탄화: 파일 업로드를 기본으로 바로 보여주고, 수기 입력은 접이식으로(탭 중첩 제거).
            with ui.column().classes("w-full gap-3"):

                # -- File upload (기본) --
                with ui.column().classes("w-full"):
                    with ui.element("div").classes("dg-banner dg-banner-info w-full mb-3"):
                        ui.icon("info", size="18px")
                        ui.label(
                            "CSV: 당근 광고관리자 내려받기 파일 (헤더 자동 매핑)  |  "
                            "XLSX: 기간|비용|노출|클릭|문의|단골|쿠폰 (1행=헤더)"
                        )

                    with ui.row().classes("gap-3 items-center"):
                        ui.upload(
                            label="파일 선택 (.csv / .xlsx)",
                            auto_upload=True,
                            on_upload=lambda e: _handle_upload(e),
                            max_file_size=50_000_000,
                        ).classes("max-w-xs dg-upload").props('accept=".csv,.xlsx"')

                        ui.button(
                            "샘플 템플릿 생성", icon="download",
                            on_click=lambda: _create_sample_excel(),
                        ).classes("dg-btn-secondary dg-btn-sm")

                    upload_spinner = ui.row().classes("w-full mt-2 items-center gap-2 hidden")
                    with upload_spinner:
                        ui.spinner("dots", size="sm")
                        upload_spinner_label = ui.label("파일을 읽고 있어요...").classes("dg-progress-text")
                    upload_summary = ui.column().classes("w-full mt-2 hidden")
                    upload_preview = ui.column().classes("w-full mt-3 hidden")

                # -- Manual input (접이식) --
                with ui.expansion("직접 수기 입력", icon="edit_note").classes("w-full dg-expansion").props("dense"):
                    manual_rows_container = ui.column().classes("w-full gap-2")
                    _manual_inputs: List[dict] = []

                    def _build_manual_row(idx: int, defaults: dict | None = None) -> dict:
                        d = defaults or _blank_row(idx)
                        with manual_rows_container:
                            with ui.row().classes("w-full gap-2 items-center"):
                                p = ui.input(value=d["period_label"]).props(
                                    'placeholder="기간" outlined dense'
                                ).classes("w-28 dg-input")
                                c = ui.input(value=str(d["cost"])).props(
                                    'placeholder="비용" outlined dense type=number'
                                ).classes("w-24 dg-input")
                                im = ui.input(value=str(d["impressions"])).props(
                                    'placeholder="노출" outlined dense type=number'
                                ).classes("w-24 dg-input")
                                cl = ui.input(value=str(d["clicks"])).props(
                                    'placeholder="클릭" outlined dense type=number'
                                ).classes("w-24 dg-input")
                                inq = ui.input(value=str(d["inquiries"])).props(
                                    'placeholder="문의" outlined dense type=number'
                                ).classes("w-24 dg-input")
                                reg = ui.input(value=str(d["regulars"])).props(
                                    'placeholder="단골" outlined dense type=number'
                                ).classes("w-20 dg-input")
                                coup = ui.input(value=str(d["coupons"])).props(
                                    'placeholder="쿠폰" outlined dense type=number'
                                ).classes("w-20 dg-input")
                        return {"p": p, "c": c, "im": im, "cl": cl, "inq": inq, "reg": reg, "coup": coup}

                    with ui.row().classes("w-full gap-1 px-1"):
                        for lbl, w in [("기간", "w-28"), ("비용(원)", "w-24"), ("노출", "w-24"),
                                       ("클릭", "w-24"), ("문의", "w-24"), ("단골", "w-20"), ("쿠폰", "w-20")]:
                            ui.label(lbl).classes(f"{w} dg-label-sm")

                    for i in range(4):
                        _manual_inputs.append(_build_manual_row(i + 1))

                    with ui.row().classes("gap-3 mt-3"):
                        ui.button(
                            "행 추가", icon="add",
                            on_click=lambda: _manual_inputs.append(
                                _build_manual_row(len(_manual_inputs) + 1)
                            ),
                        ).classes("dg-btn-secondary dg-btn-sm")
                        ui.button(
                            "데이터 적용", icon="check",
                            on_click=lambda: _apply_manual_inputs(_manual_inputs),
                        ).classes("dg-btn-primary dg-btn-sm")

        # -- KPI display --
        kpi_card = ui.card().classes("dg-card w-full hidden")
        with kpi_card:
            section_header("analytics", "핵심 지표", "전 기간 대비 증감을 함께 보여 드려요.")
            kpi_grid = ui.element("div").classes("w-full").style(
                "display: grid; grid-template-columns: repeat(auto-fill, minmax(185px, 1fr)); gap: 12px;"
            )
            with ui.expansion("세부 지표 전체 보기", icon="table_view").classes(
                "w-full dg-expansion mt-2"
            ).props("dense"):
                kpi_detail_grid = ui.element("div").classes("w-full").style(
                    "display: grid; grid-template-columns: repeat(auto-fill, minmax(155px, 1fr)); gap: 12px;"
                )

        def _delta_pct(curr: float, prev: float) -> float | None:
            if prev <= 0:
                return None
            return (curr - prev) / prev * 100

        def _show_kpi(kpi: dict) -> None:
            kpi_grid.clear()
            kpi_detail_grid.clear()
            pks = kpi.get("period_kpis", [])
            last = pks[-1] if len(pks) >= 2 else None
            prev = pks[-2] if len(pks) >= 2 else None

            def delta_of(field: str) -> float | None:
                if not last or not prev:
                    return None
                return _delta_pct(last.get(field, 0), prev.get(field, 0))

            cpr = kpi.get("cpr", 0)
            cpco = kpi.get("cp_coupon", 0)
            # (라벨, 값, 보조 정보, 전기간 대비 증감, 증가가 좋은 지표인가)
            heroes = [
                ("총 광고비", f"₩{kpi.get('total_cost',0):,}",
                 f"{len(pks)}개 기간 합계", delta_of("cost"), None),
                ("CTR", f"{kpi.get('ctr',0):.2f}%",
                 f"CPC ₩{kpi.get('cpc',0):,.0f}", delta_of("ctr"), True),
                ("클릭", f"{kpi.get('total_clicks',0):,}",
                 f"노출 {kpi.get('total_impressions',0):,}", delta_of("clicks"), True),
                ("문의", f"{kpi.get('total_inquiries',0):,}건",
                 f"1건당 ₩{kpi.get('cpa',0):,.0f}" if kpi.get("cpa", 0) > 0 else "집행 결과 없음",
                 delta_of("inquiries"), True),
                ("단골", f"{kpi.get('total_regulars',0):,}명",
                 f"1명당 ₩{cpr:,.0f}" if cpr > 0 else "집행 결과 없음",
                 delta_of("regulars"), True),
                ("쿠폰", f"{kpi.get('total_coupons',0):,}건",
                 f"1건당 ₩{cpco:,.0f}" if cpco > 0 else "집행 결과 없음",
                 delta_of("coupons"), True),
            ]
            with kpi_grid:
                for label, val, sub, delta, higher_is_better in heroes:
                    card_cls = (
                        "dg-stat-card dg-stat-card--hero"
                        if label == "총 광고비" else "dg-stat-card"
                    )
                    with ui.element("div").classes(card_cls):
                        ui.label(label).classes("dg-stat-label")
                        ui.label(val).classes("dg-stat-value")
                        ui.label(sub).classes("dg-stat-sub")
                        if delta is not None:
                            up = delta >= 0
                            arrow = "▲" if up else "▼"
                            if higher_is_better is None:
                                cls = "dg-stat-delta-neutral"
                            elif up == higher_is_better:
                                cls = "dg-stat-delta-good"
                            else:
                                cls = "dg-stat-delta-bad"
                            ui.label(f"{arrow} {abs(delta):.0f}% 전 기간 대비").classes(
                                f"dg-stat-delta {cls}"
                            )

            detail = [
                ("CPC", f"{kpi.get('cpc',0):,.0f} 원"),
                ("CPM", f"{kpi.get('cpm',0):,.0f} 원"),
                ("CPA(문의당)", f"{kpi.get('cpa',0):,.0f} 원"),
                ("CPR(단골당)", f"{kpi.get('cpr',0):,.0f} 원"),
                ("쿠폰당 비용", f"{kpi.get('cp_coupon',0):,.0f} 원"),
                ("클릭→문의", f"{kpi.get('cvr_click_inquiry',0):.1f} %"),
                ("클릭→단골", f"{kpi.get('cvr_click_regular',0):.1f} %"),
                ("문의→단골", f"{kpi.get('cvr_inquiry_regular',0):.1f} %"),
                ("문의→쿠폰", f"{kpi.get('cvr_inquiry_coupon',0):.1f} %"),
                ("단골→쿠폰", f"{kpi.get('cvr_regular_coupon',0):.1f} %"),
            ]
            with kpi_detail_grid:
                for label, val in detail:
                    with ui.element("div").classes("dg-kpi-card"):
                        ui.label(label).classes("dg-kpi-label")
                        ui.label(val).classes("dg-kpi-value")

            kpi_card.classes(remove="hidden")
            _show_funnel(kpi)
            _show_trends(kpi)
            _show_profitability(kpi)
            _show_period_efficiency(kpi)

        # -- Visualization row: 퍼널 | 수익성 (2열 벤토 배치) --
        viz_row = ui.row().classes("w-full gap-5 items-stretch").style("flex-wrap: nowrap")

        with viz_row:
            funnel_card = ui.card().classes("dg-card hidden").style(
                "flex: 7 1 0; min-width: 0"
            )
        with funnel_card:
            section_header("filter_alt", "퍼널 분석", "노출 -> 클릭 -> 문의 -> 단골 -> 쿠폰")
            funnel_container = ui.element("div").classes("w-full")

        def _show_funnel(kpi: dict) -> None:
            t_imp = kpi.get("total_impressions", 0)
            t_click = kpi.get("total_clicks", 0)
            t_inq = kpi.get("total_inquiries", 0)
            t_reg = kpi.get("total_regulars", 0)
            t_coup = kpi.get("total_coupons", 0)
            coupon_rate = (t_coup / t_click * 100) if t_click > 0 else 0.0
            # 측정된 지표만 단계로 (당근 breakdown에 전환 컬럼이 없으면 노출→클릭까지만).
            available = page_state.get("metrics_available") or FULL_METRICS
            stages, has_conv = build_funnel_stages(
                impressions=t_imp, clicks=t_click,
                ctr=kpi.get("ctr", 0), cpc=kpi.get("cpc", 0), cpm=kpi.get("cpm", 0),
                conversions={
                    "inquiries": (t_inq, kpi.get("cvr_click_inquiry", 0), kpi.get("cpa", 0)),
                    "regulars": (t_reg, kpi.get("cvr_click_regular", 0), kpi.get("cpr", 0)),
                    "coupons": (t_coup, coupon_rate, kpi.get("cp_coupon", 0)),
                },
                available=available,
            )
            render_funnel(funnel_container, stages, has_conversion_data=has_conv)
            funnel_card.classes(remove="hidden")

        # -- Trend charts (기간별 추이) — 전체 폭 --
        trend_card = ui.card().classes("dg-card w-full hidden")
        with trend_card:
            section_header("show_chart", "기간별 추이", "단계별 결과 수와 단가 변화")
            trend_row = ui.row().classes("w-full gap-4 flex-wrap")

        def _show_trends(kpi: dict) -> None:
            trend_row.clear()
            pks = kpi.get("period_kpis", [])
            if len(pks) < 2:
                trend_card.classes("hidden", remove=False)
                return

            labels = [p["label"] for p in pks]
            base = {
                "tooltip": {"trigger": "axis"},
                "legend": {"top": 0, "textStyle": {"fontSize": 11}},
                "grid": {"left": 56, "right": 56, "top": 34, "bottom": 28},
            }
            counts_options = {
                **base,
                "xAxis": {"type": "category", "data": labels, "axisLabel": {"fontSize": 11}},
                "yAxis": [
                    {"type": "value", "name": "전환", "nameTextStyle": {"fontSize": 10}},
                    {"type": "value", "name": "클릭", "nameTextStyle": {"fontSize": 10}},
                ],
                "series": [
                    {"name": "클릭", "type": "line", "yAxisIndex": 1, "smooth": True,
                     "data": [p["clicks"] for p in pks], "itemStyle": {"color": "#7295C4"}},
                    {"name": "문의", "type": "line", "smooth": True,
                     "data": [p["inquiries"] for p in pks], "itemStyle": {"color": "#E08F55"}},
                    {"name": "단골", "type": "line", "smooth": True,
                     "data": [p["regulars"] for p in pks], "itemStyle": {"color": "#6FAE8F"}},
                    {"name": "쿠폰", "type": "line", "smooth": True,
                     "data": [p["coupons"] for p in pks], "itemStyle": {"color": "#9C8AB8"}},
                ],
            }
            costs_options = {
                **base,
                "xAxis": {"type": "category", "data": labels, "axisLabel": {"fontSize": 11}},
                "yAxis": [
                    {"type": "value", "name": "문의당(원)", "nameTextStyle": {"fontSize": 10}},
                    {"type": "value", "name": "클릭당(원)", "nameTextStyle": {"fontSize": 10}},
                ],
                "series": [
                    {"name": "문의당 비용", "type": "line", "smooth": True, "connectNulls": True,
                     "data": [round(p["cpa"]) if p["cpa"] > 0 else None for p in pks],
                     "itemStyle": {"color": "#E08F55"}},
                    {"name": "클릭당 비용", "type": "line", "yAxisIndex": 1, "smooth": True, "connectNulls": True,
                     "data": [round(p["cpc"]) if p["cpc"] > 0 else None for p in pks],
                     "itemStyle": {"color": "#7295C4"}},
                ],
            }

            with trend_row:
                with ui.column().classes("flex-1 gap-1").style("min-width: 340px"):
                    ui.label("단계별 결과 수").classes("dg-label-sm")
                    ui.echart(counts_options).classes("w-full").style("height: 280px")
                with ui.column().classes("flex-1 gap-1").style("min-width: 340px"):
                    ui.label("단계별 결과 단가").classes("dg-label-sm")
                    ui.echart(costs_options).classes("w-full").style("height: 280px")
            trend_card.classes(remove="hidden")

        # -- Profitability (수익성) — 퍼널 옆 칸 --
        with viz_row:
            profit_card = ui.card().classes("dg-card hidden").style(
                "flex: 5 1 0; min-width: 0"
            )
        with profit_card:
            section_header("account_balance", "수익성 분석", "현재 CPA 대비 확장 여력을 판단해 드려요.")
            profit_container = ui.column().classes("w-full gap-3")

        def _show_profitability(kpi: dict) -> None:
            profit_container.clear()
            cpa = kpi.get("cpa", 0)
            cpr = kpi.get("cpr", 0)
            if cpa <= 0:
                return

            with profit_container:
                # CPA 기반 수익성 게이지 (사용자가 목표 CPA를 입력하지 않으므로
                # 업종 평균 벤치마크 대비로 표시 — 당근 평균 CPA ~5,000~15,000원)
                # 대신 CPA 절대값과 퍼널 효율 비율로 판단
                ui.label("문의당 비용(CPA) 수준").style(
                    "font-size: 14px; font-weight: 600; color: var(--dg-text-primary)"
                )
                with ui.row().classes("w-full items-center gap-3"):
                    ui.label(f"{cpa:,.0f}원").style(
                        "font-size: 24px; font-weight: 700; color: var(--dg-primary); min-width: 120px"
                    )
                    with ui.column().classes("flex-1 gap-1"):
                        # 게이지: CPA를 0~30,000원 범위로 시각화
                        gauge_max = max(cpa * 2, 30000)
                        gauge_pct = min(cpa / gauge_max * 100, 100)
                        gauge_cls = "dg-gauge-safe" if gauge_pct < 40 else ("dg-gauge-warning" if gauge_pct < 70 else "dg-gauge-danger")
                        with ui.element("div").classes("dg-gauge w-full"):
                            ui.element("div").classes(f"dg-gauge-fill {gauge_cls}").style(
                                f"width: {gauge_pct:.0f}%"
                            )
                        verdict = "확대 가능" if gauge_pct < 40 else ("주의 필요" if gauge_pct < 70 else "효율 개선 필요")
                        verdict_color = "var(--dg-success)" if gauge_pct < 40 else ("var(--dg-warning)" if gauge_pct < 70 else "var(--dg-error)")
                        ui.label(verdict).style(
                            f"font-size: 13px; font-weight: 600; color: {verdict_color}"
                        )

                if cpr > 0:
                    ui.separator().classes("my-1")
                    ui.label("단골당 비용(CPR) 수준").style(
                        "font-size: 14px; font-weight: 600; color: var(--dg-text-primary)"
                    )
                    with ui.row().classes("w-full items-center gap-3"):
                        ui.label(f"{cpr:,.0f}원").style(
                            "font-size: 24px; font-weight: 700; color: var(--dg-primary); min-width: 120px"
                        )
                        with ui.column().classes("flex-1 gap-1"):
                            gauge_max_r = max(cpr * 2, 50000)
                            gauge_pct_r = min(cpr / gauge_max_r * 100, 100)
                            gauge_cls_r = "dg-gauge-safe" if gauge_pct_r < 40 else ("dg-gauge-warning" if gauge_pct_r < 70 else "dg-gauge-danger")
                            with ui.element("div").classes("dg-gauge w-full"):
                                ui.element("div").classes(f"dg-gauge-fill {gauge_cls_r}").style(
                                    f"width: {gauge_pct_r:.0f}%"
                                )

            profit_card.classes(remove="hidden")

        # -- Period efficiency (기간별 효율 분석) — 전체 폭 --
        period_card = ui.card().classes("dg-card w-full hidden")
        with period_card:
            section_header("compare_arrows", "기간별 효율 분석", "기간별 성과를 비교하고 예산 재배분을 시뮬레이션해요.")
            period_container = ui.column().classes("w-full gap-3")

        def _show_period_efficiency(kpi: dict) -> None:
            period_container.clear()
            period_kpis = kpi.get("period_kpis", [])
            if len(period_kpis) < 2:
                return

            eff = set(kpi.get("efficient_periods", []))
            ineff = set(kpi.get("inefficient_periods", []))
            avg_cpa = kpi.get("cpa", 0)

            with period_container:
                # 기간별 KPI 테이블 (컴팩트 뷰)
                table_columns = [
                    {"name": "label", "label": "기간", "field": "label", "align": "left", "sortable": True},
                    {"name": "status", "label": "상태", "field": "status", "align": "center"},
                    {"name": "cost", "label": "비용", "field": "cost", "align": "right", "sortable": True},
                    {"name": "impressions", "label": "노출", "field": "impressions", "align": "right", "sortable": True},
                    {"name": "clicks", "label": "클릭", "field": "clicks", "align": "right", "sortable": True},
                    {"name": "ctr", "label": "CTR", "field": "ctr", "align": "right", "sortable": True},
                    {"name": "inquiries", "label": "문의", "field": "inquiries", "align": "right", "sortable": True},
                    {"name": "cpa", "label": "CPA", "field": "cpa", "align": "right", "sortable": True},
                ]
                table_rows = []
                for pk in period_kpis:
                    plabel = pk["label"]
                    if plabel in eff:
                        status = "효율"
                    elif plabel in ineff:
                        status = "비효율"
                    else:
                        status = "보통"
                    table_rows.append({
                        "label": plabel,
                        "status": status,
                        "cost": f"{pk['cost']:,}원",
                        "impressions": f"{pk['impressions']:,}",
                        "clicks": f"{pk['clicks']:,}",
                        "ctr": f"{pk['ctr']:.2f}%",
                        "inquiries": f"{pk['inquiries']:,}",
                        "cpa": f"{pk['cpa']:,.0f}원" if pk["cpa"] > 0 else "-",
                    })

                period_table = ui.table(
                    columns=table_columns, rows=table_rows, row_key="label",
                ).classes("w-full dg-table").props("dense flat bordered")
                period_table.add_slot("body-cell-status", r'''
                    <q-td :props="props">
                        <q-badge
                            :color="props.value === '효율' ? 'green' : props.value === '비효율' ? 'red' : 'grey'"
                            :label="props.value"
                        />
                    </q-td>
                ''')

                # 예산 재배분 시뮬레이션 배너
                extra_conv = kpi.get("realloc_extra_conversions", 0)
                cpa_improv = kpi.get("realloc_cpa_improvement", 0.0)
                if ineff and extra_conv > 0:
                    with ui.element("div").classes("dg-banner dg-banner-info w-full mt-2"):
                        ui.icon("lightbulb", size="18px")
                        with ui.column().classes("gap-1"):
                            ui.label("예산 재배분 시뮬레이션").style("font-weight: 600; font-size: 13px")
                            ui.label(
                                f"비효율 기간({', '.join(ineff)})의 예산을 "
                                f"효율 기간({', '.join(eff)})으로 이동 시:"
                            ).style("font-size: 12px")
                            ui.label(
                                f"예상 추가 전환 +{extra_conv}건, CPA {cpa_improv:.1f}% 개선"
                            ).style("font-size: 13px; font-weight: 600; color: var(--dg-success)")

                # ON/OFF 가이드
                if eff or ineff:
                    ui.separator().classes("my-2")
                    ui.label("운영 가이드").style(
                        "font-size: 14px; font-weight: 600; color: var(--dg-text-primary)"
                    )
                    guide_rows = []
                    for label in eff:
                        guide_rows.append({"기간": label, "판단": "유지/증액", "사유": f"CPA가 평균({avg_cpa:,.0f}원) 이하"})
                    for label in ineff:
                        guide_rows.append({"기간": label, "판단": "OFF 권장", "사유": f"CPA가 평균의 1.5배 초과"})
                    if guide_rows:
                        ui.table(
                            columns=[
                                {"name": "기간", "label": "기간", "field": "기간", "align": "left"},
                                {"name": "판단", "label": "판단", "field": "판단", "align": "left"},
                                {"name": "사유", "label": "사유", "field": "사유", "align": "left"},
                            ],
                            rows=guide_rows,
                        ).classes("w-full dg-table")

            period_card.classes(remove="hidden")

        # -- Chart preview --
        chart_card = ui.card().classes("dg-card w-full hidden")
        chart_row = ui.row().classes("w-full gap-4 flex-wrap")

        with chart_card:
            section_header("bar_chart", "성과 차트 미리보기")
            chart_row

        # -- AI report generation --
        with ui.card().classes("dg-card w-full"):
            section_header("smart_toy", "AI 보고서 생성", "AI가 성과 데이터를 분석해 인사이트 보고서를 써 드려요.")
            with ui.row().classes("items-start gap-8"):
                with ui.column().classes("gap-1"):
                    ui.label("AI 엔진").classes("dg-label-sm")
                    engine_radio = ui.radio(
                        {"claude": "Claude", "gpt": "GPT", "coordinate": "Claude+GPT 조율"},
                        value="claude",
                    ).props("inline").classes("dg-radio")
                with ui.column().classes("flex-1 gap-1"):
                    ui.label("추가 요청 사항 (선택)").classes("dg-label-sm")
                    extra_input = ui.textarea(
                        placeholder="예: 다음 달 예산 20% 증가 검토 중, ROI 중심으로 분석 등"
                    ).classes("w-full dg-input").props("rows=2 outlined")

        with ui.row().classes("gap-3 items-center"):
            gen_btn = ui.button(
                "보고서 생성", icon="description",
                on_click=lambda: _generate_report(),
            ).classes("dg-btn-primary")
            # 내보내기는 메뉴 하나로 묶는다(버튼 과밀 방지). 광고주용은 슬라이드를 추천 상위에.
            export_btn = ui.button("내보내기", icon="download").props(
                "icon-right=expand_more"
            ).classes("dg-btn-secondary")
            with export_btn:
                with ui.menu().props('anchor="bottom left" self="top left"'):
                    ui.menu_item("슬라이드 PPT (추천)", on_click=lambda: _export_pptx())
                    ui.menu_item("슬라이드 HTML", on_click=lambda: _export_slides())
                    ui.separator()
                    ui.menu_item("문서(DOCX) · 기본 폴더", on_click=lambda: _export_default())
                    ui.menu_item("문서(DOCX) · 다른 위치로", on_click=lambda: _export_saveas())
            cancel_btn = ui.button(
                "중단", icon="stop",
                on_click=lambda: _cancel_generation(),
            ).classes("dg-btn-danger dg-btn-sm hidden")
            spinner = ui.spinner(size="32px").classes("hidden")
            step_label = ui.label("").classes("dg-progress-text hidden")
            download_status = ui.label("").style(
                "font-size: 13px; font-weight: 600; color: var(--dg-success)"
            ).classes("hidden")

        def _cancel_generation() -> None:
            page_state["cancelled"] = True
            step_label.set_text("중단하고 있어요...")

        def _set_step(text: str) -> None:
            step_label.classes(remove="hidden")
            step_label.set_text(text)

        # -- Report preview --
        report_card = ui.card().classes("dg-card w-full hidden")
        # "이렇게 하겠다" 운영 계획 — 퍼널/지표 아래, 본문 위에 먼저 띄운다(중간 보고서의 핵심).
        action_container = ui.column().classes("w-full hidden")
        report_md = ui.markdown("").classes("w-full dg-prose")
        judgment_container = ui.column().classes("w-full hidden")

        # ── 화면 배열 재구성 ──
        # 대시보드(핵심 지표 → 퍼널|수익성 → 추이 → 기간 효율)를 데이터 입력
        # 카드보다 위로 올린다. 입력은 한 번 하고 보는 건 매일 보기 때문.
        # (0=제목, 1=부제, 2=프로젝트 카드 다음 위치로 이동)
        for _idx, _el in enumerate((kpi_card, viz_row, trend_card, period_card)):
            _el.move(target_index=3 + _idx)

        with report_card:
            section_header("summarize", "분석 결과")
            action_container
            report_md
            judgment_container

        def _render_action_plan(content: str) -> None:
            """'그래서 이렇게 하겠다' — Next Actions + 다음 실험을 본문 위에 카드로 먼저 띄운다.

            데이터(퍼널·지표)만 보고 끝나지 않도록, 중간 보고서의 핵심인 '다음 운영 계획'을
            구조화해 앞단에 노출한다. 이미 _parse_ai_insights가 뽑은 구조를 재사용한다.
            """
            insights = _parse_ai_insights(content)
            next_actions = [a for a in (insights.get("next_actions") or []) if str(a).strip()]
            experiments = insights.get("experiments") or []
            if not next_actions and not experiments:
                action_container.classes(add="hidden")
                return
            action_container.clear()
            with action_container:
                with ui.element("div").classes("dg-banner dg-banner-info w-full"):
                    ui.icon("playlist_add_check", size="18px")
                    ui.label("다음 기간 운영 계획 — 이렇게 하겠습니다").style(
                        "font-size: 15px; font-weight: 700; color: var(--dg-text-primary)"
                    )
                if next_actions:
                    with ui.column().classes("w-full gap-1").style("margin-top: 4px"):
                        for i, act in enumerate(next_actions, 1):
                            with ui.row().classes("items-start gap-2 w-full no-wrap"):
                                ui.label(str(i)).classes("dg-step-num").style(
                                    "min-width: 22px; height: 22px; border-radius: 6px;"
                                    "background: var(--dg-primary); color: white; font-size: 12px;"
                                    "font-weight: 700; display: flex; align-items: center;"
                                    "justify-content: center; flex-shrink: 0; margin-top: 2px"
                                )
                                ui.label(str(act)).style(
                                    "font-size: 14px; line-height: 1.5; color: var(--dg-text-secondary)"
                                )
                if experiments:
                    ui.label("실행 계획 (우선순위·변경·성공 기준·일정)").style(
                        "font-size: 13px; font-weight: 600; color: var(--dg-text-tertiary); margin-top: 8px"
                    )
                    ui.table(
                        columns=[
                            {"name": "priority", "label": "순위", "field": "priority", "align": "left"},
                            {"name": "change", "label": "변경 내용", "field": "change", "align": "left"},
                            {"name": "success_criteria", "label": "성공 기준", "field": "success_criteria", "align": "left"},
                            {"name": "schedule", "label": "일정", "field": "schedule", "align": "left"},
                        ],
                        rows=[
                            {
                                "priority": e.get("priority", "-"),
                                "change": e.get("change", "-"),
                                "success_criteria": e.get("success_criteria", "-"),
                                "schedule": e.get("schedule", "-"),
                            }
                            for e in experiments
                        ],
                    ).classes("w-full dg-table").props("dense flat bordered")
            action_container.classes(remove="hidden")

        def _render_judgment_table(content: str) -> None:
            insights = _parse_ai_insights(content)
            j = insights.get("judgment", {})
            if not j:
                judgment_container.classes(add="hidden")
                return
            label_map = {"expand": "확대", "review": "검토", "stop": "중단"}
            rows_data = []
            for key in ["expand", "review", "stop"]:
                if key in j:
                    rows_data.append({"판단": label_map.get(key, key), "기준": j[key]})
            if not rows_data:
                judgment_container.classes(add="hidden")
                return
            judgment_container.clear()
            with judgment_container:
                ui.label("판단기준").style(
                    "font-size: 16px; font-weight: 700; color: var(--dg-text-primary); margin-top: 16px"
                )
                ui.table(
                    columns=[
                        {"name": "판단", "label": "판단", "field": "판단", "align": "left"},
                        {"name": "기준", "label": "기준", "field": "기준", "align": "left"},
                    ],
                    rows=rows_data,
                ).classes("w-full dg-table")
            judgment_container.classes(remove="hidden")

        # -- Data handlers --

        async def _set_rows(rows: List[Dict]) -> None:
            page_state["rows"] = rows
            kpi = calc_kpi(rows)
            page_state["kpi"] = kpi
            _show_kpi(kpi)
            await _render_charts(rows)
            pid = nicegui_app.storage.user.get("current_project_id")
            if pid:
                nicegui_app.storage.user["current_project_id"] = pid
                save_performance_rows(pid, rows)
                ui.notify(f"{len(rows)}개 행을 저장했어요.", type="positive")

        def _show_upload_preview(rows: List[Dict]) -> None:
            upload_preview.clear()
            upload_preview.classes(remove="hidden")
            with upload_preview:
                ui.label(f"업로드 데이터 미리보기 ({len(rows)}행)").style(
                    "font-size: 13px; font-weight: 600; color: var(--dg-text-primary)"
                )
                columns = [
                    {"name": "period_label", "label": "기간", "field": "period_label"},
                    {"name": "cost", "label": "비용", "field": "cost"},
                    {"name": "impressions", "label": "노출", "field": "impressions"},
                    {"name": "clicks", "label": "클릭", "field": "clicks"},
                    {"name": "inquiries", "label": "문의", "field": "inquiries"},
                    {"name": "regulars", "label": "단골", "field": "regulars"},
                    {"name": "coupons", "label": "쿠폰", "field": "coupons"},
                ]
                ui.table(columns=columns, rows=rows).classes("w-full dg-table").props("dense flat bordered")

        async def _handle_upload(e) -> None:
            try:
                page_state["segment_rows"] = []  # 새 업로드마다 초기화(데모그래픽 양식만 다시 채움)
                file_bytes = await e.file.read()
                filename = e.file.name or ""
                ext = Path(filename).suffix.lower()

                upload_spinner_label.set_text(f"'{filename}' 파일을 읽고 있어요...")
                upload_spinner.classes(remove="hidden")
                upload_summary.clear()
                upload_summary.classes(remove="hidden")

                loop = asyncio.get_running_loop()

                if ext == ".csv":
                    page_state["metrics_available"] = set(FULL_METRICS)
                    csv_rows, warnings = await loop.run_in_executor(
                        None, parse_daangn_csv, file_bytes,
                    )
                    rows = [{
                        "period_label": r["date"],
                        "cost": r["cost"],
                        "impressions": r["impressions"],
                        "clicks": r["clicks"],
                        "inquiries": r["inquiries"],
                        "regulars": r["regulars"],
                        "coupons": r["coupons"],
                    } for r in csv_rows]
                    with upload_summary:
                        with ui.element("div").classes("dg-banner dg-banner-success w-full"):
                            ui.icon("check_circle", size="18px")
                            with ui.column().classes("gap-0"):
                                ui.label(f"CSV 파싱 결과: {len(rows)}행 매핑됨")
                                mapped_cols = [
                                    k for k in ("date", "cost", "impressions", "clicks",
                                                "inquiries", "regulars", "coupons")
                                    if csv_rows and k in csv_rows[0]
                                ]
                                ui.label(f"매핑 컬럼: {', '.join(mapped_cols)}").style("font-size: 12px; opacity: 0.8")
                                if warnings:
                                    ui.label(f"경고 {len(warnings)}건").style("font-size: 12px; opacity: 0.8")
                    if not rows:
                        ui.notify("CSV에서 분석할 데이터를 찾지 못했어요. 파일 양식을 확인해 주세요.", type="warning")
                        return
                elif ext == ".xlsx":
                    # 1순위: 당근 광고관리자 직접 내보내기(long-format) 자동 인식 → 일자별 롤업
                    breakdown = await loop.run_in_executor(
                        None, _rows_from_daangn_breakdown, file_bytes,
                    )
                    if breakdown is not None:
                        rows, info = breakdown
                        warning = None
                        avail = set(info.get("metrics_available", []))
                        page_state["metrics_available"] = avail
                        page_state["segment_rows"] = info.get("segment_rows", [])
                        missing = [
                            name for key, name in (
                                ("inquiries", "문의"), ("regulars", "단골"), ("coupons", "쿠폰"),
                            ) if key not in avail
                        ]
                        with upload_summary:
                            with ui.element("div").classes("dg-banner dg-banner-success w-full"):
                                ui.icon("check_circle", size="18px")
                                with ui.column().classes("gap-0"):
                                    ui.label(
                                        f"당근 광고관리자 양식 인식 — {len(rows)}일치를 날짜별로 집계했어요."
                                    )
                                    camps = info.get("campaign_names", [])
                                    if camps:
                                        ui.label(f"캠페인 {len(camps)}개 합산").style(
                                            "font-size: 12px; opacity: 0.8")
                                    if missing:
                                        ui.label(
                                            f"이 파일엔 {'·'.join(missing)} 전환 데이터가 없어요 — "
                                            f"노출·클릭·비용 중심으로 보여 드려요. "
                                            f"당근 내보내기에서 해당 항목을 추가하면 더 깊게 분석돼요."
                                        ).style("font-size: 12px; opacity: 0.8")
                    else:
                        # 2순위: 기간|비용|노출|클릭|문의|단골|쿠폰 7열 템플릿
                        page_state["metrics_available"] = set(FULL_METRICS)
                        rows, warning = await loop.run_in_executor(
                            None, _parse_excel, file_bytes,
                        )
                        with upload_summary:
                            with ui.element("div").classes("dg-banner dg-banner-success w-full"):
                                ui.icon("check_circle", size="18px")
                                with ui.column().classes("gap-0"):
                                    ui.label(f"XLSX 파싱 결과: {len(rows)}행 로드됨")
                                    if warning:
                                        ui.label(f"경고: {warning}").style("font-size: 12px; opacity: 0.8")
                        if warning:
                            ui.notify(f"경고: {warning}", type="warning", timeout=8000)
                    if not rows:
                        ui.notify("파일에서 데이터를 찾지 못했어요. 양식을 확인해 주세요.", type="warning")
                        return
                else:
                    ui.notify(f"{ext} 형식은 지원하지 않아요. CSV 또는 XLSX 파일로 올려 주세요.", type="negative")
                    return

                upload_spinner.classes("hidden")
                upload_summary.classes(remove="hidden")
                await _set_rows(rows)
                _show_upload_preview(rows)
            except Exception as exc:
                upload_spinner.classes("hidden")
                ui.notify(f"파일을 읽지 못했어요. 양식을 확인하고 다시 올려 주세요. ({exc})", type="negative")

        async def _apply_manual_inputs(inputs: List[dict]) -> None:
            rows = []
            for inp in inputs:
                rows.append({
                    "period_label": inp["p"].value.strip() or f"기간{len(rows)+1}",
                    "cost": _parse_int(inp["c"].value),
                    "impressions": _parse_int(inp["im"].value),
                    "clicks": _parse_int(inp["cl"].value),
                    "inquiries": _parse_int(inp["inq"].value),
                    "regulars": _parse_int(inp["reg"].value),
                    "coupons": _parse_int(inp["coup"].value),
                })
            rows = [r for r in rows if any(r[k] > 0 for k in ("cost", "impressions", "clicks", "inquiries"))]
            if not rows:
                ui.notify("분석할 데이터가 아직 없어요. 행을 입력하거나 파일을 올려 주세요.", type="warning")
                return
            page_state["metrics_available"] = set(FULL_METRICS)
            await _set_rows(rows)

        async def _render_charts(rows: List[Dict]) -> None:
            chart_row.clear()
            loop = asyncio.get_running_loop()
            paths = await loop.run_in_executor(None, make_charts, rows, CHARTS_DIR)
            if not paths:
                return
            chart_card.classes(remove="hidden")
            with chart_row:
                for p in paths:
                    if p.exists():
                        ui.image(str(p)).classes("dg-chart-img")

        async def _load_saved_data() -> None:
            pid = nicegui_app.storage.user.get("current_project_id")
            if not pid:
                return

            # 이전 데이터 초기화
            page_state["rows"] = []
            page_state["kpi"] = {}
            page_state["report_content"] = ""
            page_state["metrics_available"] = set(FULL_METRICS)
            page_state["c_text"] = ""
            page_state["g_text"] = ""
            page_state["engine"] = "claude"
            kpi_grid.clear()
            kpi_detail_grid.clear()
            kpi_card.classes("hidden")
            funnel_container.clear()
            funnel_card.classes("hidden")
            trend_row.clear()
            trend_card.classes("hidden")
            profit_container.clear()
            profit_card.classes("hidden")
            chart_row.clear()
            chart_card.classes("hidden")
            period_container.clear()
            period_card.classes("hidden")
            report_md.set_content("")
            action_container.classes(add="hidden")
            report_card.classes("hidden")
            upload_preview.clear()
            upload_preview.classes("hidden")

            # 새 프로젝트 데이터 로드
            rows = get_performance_rows(pid)
            if rows:
                page_state["rows"] = rows
                kpi = calc_kpi(rows)
                page_state["kpi"] = kpi
                _show_kpi(kpi)
                await _render_charts(rows)
            rpt = get_latest_report(pid)
            if rpt:
                page_state["report_content"] = rpt["content"]
                report_md.set_content(rpt["content"])
                _render_action_plan(rpt["content"])
                _render_judgment_table(rpt["content"])
                report_card.classes(remove="hidden")

        async def _generate_report() -> None:
            pid = nicegui_app.storage.user.get("current_project_id")
            if not pid:
                ui.notify("프로젝트를 먼저 선택해 주세요.", type="warning")
                return
            # storage 동기화
            nicegui_app.storage.user["current_project_id"] = pid
            rows = page_state.get("rows", [])
            if not rows:
                ui.notify("성과 데이터를 먼저 입력하거나 올려 주세요.", type="warning")
                return

            project = get_project(pid)
            if not project:
                ui.notify("프로젝트를 찾을 수 없어요. 프로젝트 페이지에서 다시 선택해 주세요.", type="negative")
                return

            engine = engine_radio.value
            extra = extra_input.value
            kpi = page_state.get("kpi", calc_kpi(rows))

            page_state["cancelled"] = False
            spinner.classes(remove="hidden")
            gen_btn.props("disabled")
            cancel_btn.classes(remove="hidden")

            try:
                _set_step("1/4 보고서 작성을 준비하고 있어요...")
                prompt = build_report_prompt(
                    project, rows, kpi, extra,
                    wiki=store_wiki.wiki_context(pid, project, kpi),
                )
                loop = asyncio.get_running_loop()

                if page_state["cancelled"]:
                    ui.notify("생성을 중단했어요.", type="warning")
                    return

                guide = SYSTEM_GUIDE_REPORT + knowledge.domain_knowledge("report")
                if engine == "coordinate":
                    from app.ai.coordination import synthesize
                    _set_step("2/4 Claude와 GPT가 각자 초안을 쓰고 있어요...")
                    claude_p = get_provider("claude")
                    gpt_p = get_provider("gpt")
                    c_text, g_text = await asyncio.gather(
                        loop.run_in_executor(None, lambda: claude_p.generate_text(prompt, system_prompt=guide)),
                        loop.run_in_executor(None, lambda: gpt_p.generate_text(prompt, system_prompt=guide)),
                    )
                    if page_state["cancelled"]:
                        ui.notify("생성을 중단했어요.", type="warning")
                        return
                    _set_step("3/4 Claude가 두 초안을 종합하고 있어요...")
                    content = await loop.run_in_executor(
                        None, lambda: synthesize(c_text, g_text, "성과 분석 보고서"),
                    )
                    page_state["c_text"] = ""
                    page_state["g_text"] = ""
                else:
                    engine_name = "GPT" if engine == "gpt" else "Claude"
                    _set_step(f"2/4 {engine_name}가 보고서를 작성하고 있어요...")
                    provider = get_provider(engine)
                    content = await loop.run_in_executor(None, lambda: provider.generate_text(prompt, system_prompt=guide))
                    if page_state["cancelled"]:
                        ui.notify("생성을 중단했어요.", type="warning")
                        return
                    page_state["c_text"] = ""
                    page_state["g_text"] = ""

                _set_step("3/4 결과를 저장하고 있어요...")
                page_state["report_content"] = content
                page_state["engine"] = engine
                save_report_content(pid, engine, content)
                # 누적 루프: 이번 분석의 병목·검증 패턴을 매장 위키에 자동 반영
                try:
                    _period = (
                        f"{rows[0]['period_label']}~{rows[-1]['period_label']}" if rows else ""
                    )
                    store_wiki.update_wiki_from_report(pid, project, kpi, content, period=_period)
                except Exception:
                    _report_log.exception("매장 위키 자동 갱신 실패 (보고서는 정상 저장됨)")
                report_md.set_content(content)
                _render_action_plan(content)
                _render_judgment_table(content)
                report_card.classes(remove="hidden")

                try:
                    _set_step("4/4 DOCX 파일을 만들고 있어요...")
                    project_name = project.get('name', 'report')
                    download_status.classes(remove="hidden")
                    download_status.set_text("DOCX 파일을 준비하고 있어요...")
                    if engine == "both":
                        c_bytes, g_bytes = await asyncio.wait_for(
                            asyncio.gather(
                                loop.run_in_executor(None, _make_report_docx_bytes, project, rows, kpi, c_text),
                                loop.run_in_executor(None, _make_report_docx_bytes, project, rows, kpi, g_text),
                            ),
                            timeout=90.0,
                        )
                        c_fname = f"성과보고서_{project_name}_Claude.docx"
                        g_fname = f"성과보고서_{project_name}_Gemini.docx"
                        ExportManager.save_default(c_bytes, filename=c_fname)
                        ExportManager.save_default(g_bytes, filename=g_fname)
                        download_status.set_text(f"{c_fname}, {g_fname} 다운로드를 시작했어요")
                        ui.notify(
                            f"보고서가 완성됐어요!\n{c_fname}\n{g_fname}",
                            type="positive", timeout=10000, close_button="확인",
                        )
                    else:
                        docx_bytes = await asyncio.wait_for(
                            loop.run_in_executor(
                                None, _make_report_docx_bytes, project, rows, kpi, content
                            ),
                            timeout=90.0,
                        )
                        fname = f"성과보고서_{project_name}.docx"
                        ExportManager.save_default(docx_bytes, filename=fname)
                        download_status.set_text(f"{fname} 다운로드를 시작했어요")
                        ui.notify(
                            f"보고서가 완성됐어요!\n{fname}",
                            type="positive", timeout=8000, close_button="확인",
                        )
                except asyncio.TimeoutError:
                    download_status.set_text("DOCX 만들기가 너무 오래 걸려 중단했어요")
                    ui.notify(
                        "DOCX 만들기가 90초를 넘겨 중단했어요. 보고서 텍스트는 저장해 두었어요.",
                        type="warning", timeout=10000, close_button="확인",
                    )
                except Exception as docx_err:
                    download_status.set_text("DOCX 파일을 만들지 못했어요")
                    ui.notify(f"보고서는 완성했지만 DOCX 파일은 만들지 못했어요. ({docx_err})", type="warning", timeout=8000)

            except Exception as exc:
                ui.notify(f"보고서를 만들지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative", timeout=8000)
            finally:
                spinner.classes("hidden")
                cancel_btn.classes("hidden")
                step_label.classes("hidden")
                gen_btn.props(remove="disabled")

        def _validate_export() -> tuple:
            content = page_state.get("report_content", "")
            if not content:
                raise ValueError("보고서를 먼저 만들어 주세요.")
            rows = page_state.get("rows", [])
            kpi = page_state.get("kpi", {})
            pid = nicegui_app.storage.user.get("current_project_id")
            project = get_project(pid) if pid else None
            if not project:
                raise ValueError("프로젝트를 선택해 주세요.")
            project_name = project.get('name', 'report')
            engine = page_state.get("engine", "claude")
            return project, project_name, rows, kpi, content, engine

        async def _build_report_pairs() -> list[tuple[bytes, str]]:
            project, project_name, rows, kpi, content, engine = _validate_export()
            loop = asyncio.get_running_loop()
            try:
                if engine == "both" and page_state.get("c_text") and page_state.get("g_text"):
                    c_bytes, g_bytes = await asyncio.wait_for(
                        asyncio.gather(
                            loop.run_in_executor(None, _make_report_docx_bytes, project, rows, kpi, page_state["c_text"]),
                            loop.run_in_executor(None, _make_report_docx_bytes, project, rows, kpi, page_state["g_text"]),
                        ),
                        timeout=90.0,
                    )
                    return [
                        (c_bytes, f"성과보고서_{project_name}_Claude.docx"),
                        (g_bytes, f"성과보고서_{project_name}_Gemini.docx"),
                    ]
                else:
                    docx_bytes = await asyncio.wait_for(
                        loop.run_in_executor(
                            None, _make_report_docx_bytes, project, rows, kpi, content
                        ),
                        timeout=90.0,
                    )
                    return [(docx_bytes, f"성과보고서_{project_name}.docx")]
            except asyncio.TimeoutError:
                raise ValueError("DOCX 만들기가 90초를 넘겨 중단했어요. 데이터 양을 줄여서 다시 시도해 주세요.")

        async def _export_default() -> None:
            export_btn.props("disabled loading")
            download_status.classes(remove="hidden")
            download_status.set_text("DOCX 파일을 준비하고 있어요...")
            try:
                pairs = await _build_report_pairs()
                names = []
                for docx_bytes, fname in pairs:
                    ExportManager.save_default(docx_bytes, filename=fname)
                    names.append(fname)
                download_status.set_text(f"저장했어요: {', '.join(names)}")
                ui.notify(
                    "\n".join(f"{n}" for n in names),
                    type="positive", timeout=8000, close_button="확인",
                )
            except ValueError as ve:
                ui.notify(str(ve), type="warning")
            except Exception as exc:
                download_status.set_text("내보내지 못했어요")
                ui.notify(f"파일을 내보내지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative")
            finally:
                export_btn.props(remove="disabled loading")

        async def _export_saveas() -> None:
            export_btn.props("disabled loading")
            download_status.classes(remove="hidden")
            download_status.set_text("DOCX 파일을 준비하고 있어요...")
            try:
                pairs = await _build_report_pairs()
                ok = await ExportManager.save_as_multi(pairs)
                if ok:
                    names = [f for _, f in pairs]
                    download_status.set_text(f"저장했어요: {', '.join(names)}")
                else:
                    download_status.set_text("저장 취소됨")
            except ValueError as ve:
                ui.notify(str(ve), type="warning")
            except Exception as exc:
                download_status.set_text("내보내지 못했어요")
                ui.notify(f"파일을 내보내지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative")
            finally:
                export_btn.props(remove="disabled loading")

        def _gather_slide_data() -> tuple:
            """슬라이드(HTML/PPTX) 공용 데이터 수집 → (meta, kpi, insights, stages, safe_name)."""
            content = page_state.get("report_content") or ""
            kpi = page_state.get("kpi") or {}
            if not content.strip():
                raise ValueError("먼저 'AI 보고서 생성'을 눌러 보고서를 만들어 주세요. (액션 슬라이드에 필요해요)")
            pid = nicegui_app.storage.user.get("current_project_id")
            project = get_project(pid) if pid else {}
            available = page_state.get("metrics_available") or FULL_METRICS
            t_click = kpi.get("total_clicks", 0)
            t_coup = kpi.get("total_coupons", 0)
            coupon_rate = (t_coup / t_click * 100) if t_click > 0 else 0.0
            stages, _has = build_funnel_stages(
                impressions=kpi.get("total_impressions", 0), clicks=t_click,
                ctr=kpi.get("ctr", 0), cpc=kpi.get("cpc", 0), cpm=kpi.get("cpm", 0),
                conversions={
                    "inquiries": (kpi.get("total_inquiries", 0), kpi.get("cvr_click_inquiry", 0), kpi.get("cpa", 0)),
                    "regulars": (kpi.get("total_regulars", 0), kpi.get("cvr_click_regular", 0), kpi.get("cpr", 0)),
                    "coupons": (t_coup, coupon_rate, kpi.get("cp_coupon", 0)),
                },
                available=available,
            )
            meta = {
                "name": (project or {}).get("name", ""),
                "campaign_name": (project or {}).get("campaign_name", ""),
                "region": (project or {}).get("region", ""),
                "period": (project or {}).get("period", ""),
            }
            safe = sanitize_filename(meta["name"] or "보고서")
            segs = page_state.get("segment_rows") or []
            return meta, kpi, _parse_ai_insights(content), stages, segs, safe

        def _build_slides_bytes() -> tuple[bytes, str]:
            """성과 데이터+AI 보고서를 슬라이드 HTML로. (bytes, filename)."""
            from app.reporting.slides_html import build_slides_html
            from datetime import datetime
            meta, kpi, insights, stages, segs, safe = _gather_slide_data()
            html = build_slides_html(meta, kpi, insights, funnel_stages=stages,
                                     segment_rows=segs,
                                     generated_at=datetime.now().strftime("%Y-%m-%d"))
            return html.encode("utf-8"), f"{safe}_성과보고서_슬라이드_{datetime.now().strftime('%Y%m%d')}.html"

        def _build_pptx_bytes() -> tuple[bytes, str]:
            """성과 데이터+AI 보고서를 편집 가능한 PPTX로. (bytes, filename)."""
            from app.reporting.slides_pptx import build_slides_pptx
            from datetime import datetime
            meta, kpi, insights, stages, segs, safe = _gather_slide_data()
            data = build_slides_pptx(meta, kpi, insights, funnel_stages=stages,
                                     segment_rows=segs,
                                     generated_at=datetime.now().strftime("%Y-%m-%d"))
            return data, f"{safe}_성과보고서_슬라이드_{datetime.now().strftime('%Y%m%d')}.pptx"

        async def _do_slide_export(builder, btn) -> None:
            btn.props("disabled loading")
            download_status.classes(remove="hidden")
            download_status.set_text("슬라이드 보고서를 만들고 있어요...")
            try:
                loop = asyncio.get_running_loop()
                data, fname = await loop.run_in_executor(None, builder)
                ExportManager.save_default(data, filename=fname)
                download_status.set_text(f"저장했어요: {fname}")
                hint = "\n브라우저로 열고 Ctrl+P로 PDF 저장도 돼요." if fname.endswith(".html") else "\n파워포인트에서 열어 자유롭게 편집하세요."
                ui.notify(f"{fname}{hint}", type="positive", timeout=8000, close_button="확인")
            except ValueError as ve:
                ui.notify(str(ve), type="warning")
            except Exception as exc:
                download_status.set_text("내보내지 못했어요")
                ui.notify(f"슬라이드를 만들지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative")
            finally:
                btn.props(remove="disabled loading")

        async def _export_slides() -> None:
            await _do_slide_export(_build_slides_bytes, export_btn)

        async def _export_pptx() -> None:
            await _do_slide_export(_build_pptx_bytes, export_btn)

        # initial load -- use background_tasks to schedule async init
        import nicegui
        nicegui.background_tasks.create(_load_saved_data())


@ui.page("/report")
def report_page() -> None:
    """성과 보고서 + 고급 분석 통합 페이지 (완전 병합 Stage A).

    같은 당근 데이터로 [성과 요약 | 세그먼트 심화] 두 탭을 한 화면에서 본다.
    /analysis 단독 라우트는 이 페이지로 리다이렉트된다.
    """
    create_nav("/report")
    from app.pages.analysis import render_analysis_body
    with ui.tabs().classes("w-full dg-tabs").props("align=left active-color=primary indicator-color=primary") as _tabs:
        _t_perf = ui.tab("성과 요약")
        _t_seg = ui.tab("세그먼트 심화")
    with ui.tab_panels(_tabs, value=_t_perf).classes("w-full"):
        with ui.tab_panel(_t_perf).classes("p-0"):
            render_report_body()
        with ui.tab_panel(_t_seg).classes("p-0"):
            render_analysis_body()
