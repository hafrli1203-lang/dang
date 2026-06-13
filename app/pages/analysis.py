"""Screen 4 — 고급 분석 (연령/성별 찢기 + 캠페인 판정 + 재배분 시뮬 + AI 보고서).

Flow:
  1. xlsx 업로드 → 파싱 미리보기
  2. '분석 실행' 버튼 → 규칙 기반 판정 + AI 자연어 해석 (Claude/Gemini)
  3. 결과: 한 줄 요약 → 현황 진단 → 개선점 → 실행 계획 → 예상 효과
  4. '보고서 DOCX 다운로드' 버튼 → 광고주 전달용 문서 생성

Operator playbook: 당근은 머신러닝 부재 → 연령·성별 직접 찢기, 수동+자동 페어 운영.
"""
from __future__ import annotations

import asyncio
import logging
from pathlib import Path
from typing import Any

from nicegui import ui, app as nicegui_app

from app.common import create_nav
from app.database import get_project, get_projects
from app.export_manager import ExportManager
from app.paths import CHARTS_DIR, EXPORTS_DIR
from app.reporting.demographic import (
    analyze_segments,
    build_priority_checklist,
    calc_economics,
    calc_funnel,
    check_auto_manual_pairing,
    check_variable_control,
    group_ages_by_cpa,
    judge_campaigns,
    parse_demographic_xlsx,
    simulate_reallocation,
)
from app.engine.revision_table import (
    REVISION_TABLE_COLUMNS,
    build_campaign_revision_table,
    revision_rows_for_table,
    revision_table_markdown,
)
from app.reporting.analysis_docx import build_analysis_docx
from app.ai_engine import (
    SYSTEM_GUIDE_ANALYSIS,
    build_analysis_prompt,
    parse_analysis_sections,
)
from app.ai.providers import get_provider, OpenAIProvider
from app.theme import section_header

_log = logging.getLogger("analysis")

_VERDICT_BADGE = {
    "캠페인OFF":    ("error",   "OFF"),
    "소재전면교체": ("warning", "소재 교체"),
    "소재정리후유지": ("success", "유지(주력)"),
    "증액":         ("primary", "증액"),
    "유지":         ("info",    "유지"),
}


def _fmt_won(v: float | int) -> str:
    return f"{int(round(v)):,}원"


def _strip_filename(fn: str) -> str:
    """Drop suffix and date-suffix-style tokens from xlsx filename for a clean title."""
    name = Path(fn).stem if fn else ""
    # Remove trailing _YYYYMMDD_YYYYMMDD or similar
    import re
    name = re.sub(r"_\d{8}_\d{8}$", "", name)
    name = re.sub(r"_\d{8}$", "", name)
    return name.strip()


def _derive_meta_from_file(state: dict) -> dict:
    """Best-effort extraction of 광고주/지역/캠페인/기간 from uploaded xlsx alone.

    Priority order:
      - 광고주 / 캠페인: filename prefix (e.g. 윤익광고주 _ 4월다초점)
      - 캠페인 이름들의 공통 prefix (creative_key) = 캠페인
      - 지역: 캠페인명 첫 토큰 (e.g. "밀양", "유성궁동")
      - 기간: parsed meta period_first ~ period_last
    """
    parsed = state.get("parsed") or {}
    meta = parsed.get("meta") if isinstance(parsed, dict) else {} or {}
    filename = state.get("filename", "")

    camp_names = (meta.get("campaign_names") if isinstance(meta, dict) else None) or [
        c.name for c in parsed.get("campaigns", [])
    ]

    # Filename-based name/campaign split: "광고주명_캠페인_…" 형태일 때 split('_', 1)
    clean_fn = _strip_filename(filename)
    advertiser = ""
    campaign_label = ""
    if clean_fn:
        if "_" in clean_fn:
            head, rest = clean_fn.split("_", 1)
            advertiser = head.strip()
            campaign_label = rest.strip()
        else:
            advertiser = clean_fn

    # If campaign names share a common 'creative_id' prefix (no _수동/_자동), prefer it
    if camp_names and not campaign_label:
        from app.reporting.demographic import _campaign_creative_key
        creative_keys = {_campaign_creative_key(n) for n in camp_names}
        if len(creative_keys) == 1:
            campaign_label = next(iter(creative_keys))

    # Region: first underscore-separated token of the first campaign name
    region = ""
    if camp_names:
        first_token = camp_names[0].split("_", 1)[0].strip()
        # Heuristic: very short Korean prefix tends to be 지역
        if 0 < len(first_token) <= 8:
            region = first_token

    pf = meta.get("period_first", "") if isinstance(meta, dict) else ""
    pl = meta.get("period_last", "") if isinstance(meta, dict) else ""
    if pf and pl and pf != pl:
        period = f"{pf} ~ {pl}"
    elif pf or pl:
        period = pf or pl
    else:
        period = ""

    return {
        "name": advertiser or "(미지정)",
        "campaign_name": campaign_label,
        "region": region,
        "period": period,
        "industry": "",
        "author": "당근 광고 도우미",
    }


def _build_project_meta(project: dict, state: dict) -> dict:
    """Merge selected project metadata with file-derived defaults.

    Project fields take precedence when present; missing fields fall back to
    file-derived values so that 빈 프로젝트 + 파일만 있어도 일관된 표지가 나옵니다.
    """
    derived = _derive_meta_from_file(state)
    return {
        "name": project.get("name") or derived["name"],
        "region": project.get("region") or derived["region"],
        "industry": project.get("industry") or derived["industry"],
        "campaign_name": project.get("campaign_name") or derived["campaign_name"],
        "period": project.get("period") or derived["period"],
        "author": project.get("author") or derived["author"],
    }


# ───────────────────────── inspection helpers ─────────────────────────

def _inspect_xlsx(content: bytes) -> list[dict[str, Any]]:
    import io, openpyxl
    info: list[dict[str, Any]] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return info
    for sn in wb.sheetnames:
        ws = wb[sn]
        header: tuple = ()
        for row in ws.iter_rows(values_only=True, max_row=15):
            if row and sum(1 for c in row if c not in (None, "")) >= 2:
                header = row
                break
        info.append({"name": sn, "headers": [str(c) for c in header if c not in (None, "")][:12]})
    return info


# ───────────────────────── sample template ─────────────────────────

def _create_sample_xlsx() -> None:
    try:
        import openpyxl
    except ImportError:
        ui.notify("샘플 생성에 openpyxl 설치가 필요해요. 터미널에서 pip install openpyxl을 실행해 주세요.", type="negative")
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "breakdown"
        ws.append([
            "기간", "캠페인 이름", "캠페인 ID", "연령",
            "비용 (VAT 포함)", "노출 수", "도달 수", "클릭 수", "클릭률",
            "클릭당 비용(CPC)", "노출당 비용(CPM)",
            "단골 수", "후기 수", "쿠폰 다운로드 수", "관심 수", "댓글 수",
            "전화 문의 수", "채팅 문의 수", "포장 주문 수",
        ])
        rows = [
            ("2026.05.09", "A_수동", 1, "40-44", 1452, 355, 314, 6, 1.69, 242, 4090, 0, 0, 0, 0, 0, 0, 1, 0),
            ("2026.05.09", "A_수동", 1, "50-54", 1936, 539, 465, 8, 1.48, 242, 3591, 2, 0, 2, 0, 0, 0, 0, 0),
            ("2026.05.09", "A_자동", 2, "40-44", 1746, 454, 376, 6, 1.32, 291, 3845, 0, 0, 0, 0, 0, 0, 0, 0),
            ("2026.05.09", "A_자동", 2, "50-54", 2216, 595, 487, 8, 1.34, 277, 3724, 3, 0, 3, 1, 0, 0, 0, 0),
        ]
        for r in rows:
            ws.append(r)
        downloads = Path.home() / "Downloads"
        out_dir = downloads if downloads.exists() else EXPORTS_DIR
        out = out_dir / "당근광고_고급분석_샘플.xlsx"
        wb.save(out)
        ui.notify(f"샘플을 저장했어요: {out}", type="positive")
    except Exception as exc:  # noqa: BLE001
        ui.notify(f"샘플을 만들지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative")


# ───────────────────────── page ─────────────────────────

@ui.page("/analysis")
def analysis_page() -> None:
    create_nav("/analysis")

    state: dict = {
        "parsed": None,        # {"genders","ages","campaigns"}
        "judgments": None,
        "plan": None,
        "priority": None,
        "var_warnings": None,
        "pair_gaps": None,
        "ai_sections": None,
        "ai_raw": "",
        "filename": "",
    }

    with ui.column().classes("dg-page-content w-full gap-5"):

        # Page title
        ui.label("고급 분석").classes("dg-page-title")
        ui.label(
            "당근 광고관리자의 연령×캠페인 breakdown 엑셀을 분석하여, "
            "어떤 캠페인을 끄고 늘릴지부터 광고주 전달용 보고서까지 한 번에 만들어 드려요."
        ).classes("dg-page-subtitle")

        # What this page does
        with ui.card().classes("dg-card w-full"):
            section_header("insights", "이 페이지가 하는 일",
                           "분석 → 진단 → 개선점 → 실행 계획 → 광고주 보고서까지")
            with ui.row().classes("w-full gap-3 flex-wrap"):
                for icon, title, body in [
                    ("filter_list", "1. 데이터 분리",
                     "연령대별 / 캠페인별로 비용·CPA·CTR·행동을 자동 집계"),
                    ("rule", "2. 규칙 판정",
                     "OFF / 증액 / 소재교체 / 유지 — 캠페인별 4단계 판정"),
                    ("smart_toy", "3. AI 해석",
                     "Claude/Gemini가 광고주 친화 문장으로 진단·개선점 작성"),
                    ("description", "4. DOCX 출력",
                     "광고주 전달용 보고서 (표지+차트+표+멘트) 다운로드"),
                ]:
                    with ui.card().classes("dg-kpi-card").style("flex:1; min-width:200px"):
                        ui.icon(icon, size="22px").style("color: var(--dg-primary)")
                        ui.label(title).style("font-weight:700; font-size:14px; margin-top:4px")
                        ui.label(body).style("font-size:12px; color: var(--dg-text-secondary); line-height:1.4")

        # Project selector
        with ui.card().classes("dg-card w-full"):
            with ui.row().classes("items-center gap-4"):
                ui.icon("business", size="20px").style("color: var(--dg-primary)")
                ui.label("프로젝트").style("font-weight: 600")
                projects = get_projects()
                options = {p["id"]: f"{p['name']} | {p.get('region','')}" for p in projects}
                # 분석 페이지는 파일 내용 기반이라 다른 페이지의 저장된 프로젝트를 자동 회수하지 않습니다.
                # (사용자가 빈칸으로 두면 캠페인명·날짜에서 메타를 자동 추출함)
                project_sel = ui.select(
                    options, label="프로젝트 선택 (선택사항 — 빈칸이면 파일에서 자동 추출)",
                    value=None,
                ).classes("flex-1 dg-select")

        # Step 1: Upload
        with ui.card().classes("dg-card w-full"):
            section_header("upload_file", "1단계: 데이터 업로드",
                           "당근 광고관리자 → 연령/캠페인 breakdown xlsx 내보내기")

            with ui.element("div").classes("dg-banner dg-banner-info w-full mb-3"):
                ui.icon("info", size="18px")
                ui.label(
                    "받는 양식: [기간, 캠페인 이름, 연령, 비용, 노출, 클릭, 단골/문의/쿠폰...] 컬럼이 있는 1개 시트. "
                    "캠페인 이름 끝에 _수동 / _자동을 붙이면 수동/자동을 자동으로 구분해 드려요."
                )

            with ui.row().classes("w-full gap-3 items-center"):
                upload = ui.upload(
                    label="xlsx 파일 선택",
                    auto_upload=True,
                    on_upload=lambda e: _handle_upload(e),
                    max_file_size=10_000_000,
                ).classes("flex-1").props("accept=.xlsx")
                ui.button(
                    "샘플 템플릿", icon="download",
                    on_click=lambda: _create_sample_xlsx(),
                ).classes("dg-btn-secondary dg-btn-sm")

            upload_status = ui.column().classes("w-full mt-2")

        # Step 2: Run analysis
        with ui.card().classes("dg-card w-full"):
            section_header("auto_awesome", "2단계: 분석 실행",
                           "규칙 기반 판정 + AI 자연어 해석 (10~30초)")

            with ui.row().classes("w-full gap-3 items-end flex-wrap"):
                with ui.column().classes("gap-1"):
                    ui.label("AI 엔진").classes("dg-label-sm")
                    engine_sel = ui.radio(
                        {"claude": "Claude", "gpt": "GPT", "coordinate": "Claude+GPT 조율"}, value="claude",
                    ).props("inline").classes("dg-radio")
                with ui.column().classes("gap-1"):
                    ui.label("객단가 (원, 광고주 입력)").classes("dg-label-sm")
                    aov_input = ui.number(
                        value=0, min=0, step=10_000, format="%d",
                        placeholder="예: 200000",
                    ).props("outlined dense").classes("w-36 dg-input")
                with ui.column().classes("gap-1"):
                    ui.label("목표 이익률 (%, 광고주 입력)").classes("dg-label-sm")
                    margin_input = ui.number(
                        value=0, min=0, max=100, step=5, format="%d",
                        placeholder="예: 30",
                    ).props("outlined dense").classes("w-36 dg-input")
                with ui.column().classes("flex-1 gap-1"):
                    ui.label("추가 요청 (선택사항 — 비워 두셔도 돼요)").classes("dg-label-sm")
                    extra_input = ui.textarea(
                        placeholder="예: 다음 달 예산 20% 늘릴 계획 / 4050 타겟 위주 분석 / 단골 전환 중심 평가",
                    ).props("rows=2 outlined").classes("w-full dg-input")
                run_btn = ui.button(
                    "분석 실행", icon="play_arrow",
                    on_click=lambda: _run_analysis(),
                ).classes("dg-btn-primary").props("disabled")

            ui.label(
                "객단가/이익률을 입력하면 MAX CPA(허용 광고비 한계) + 한계 소진율 + "
                "손익 판정이 보고서에 추가돼요. 모르면 비워 둬도 분석은 정상으로 진행돼요."
            ).style("font-size:11px; color: var(--dg-text-tertiary); margin-top:6px")

            # ── 광고 소재 입력 (선택) ──
            ui.separator().classes("my-3")
            ui.label("광고 소재 (선택사항)").style(
                "font-size:13px; font-weight:600; color: var(--dg-text-primary)"
            )
            ui.label(
                "현재 운영 중인 소식글 본문과 썸네일을 첨부하면, AI가 카피·이미지까지 "
                "직접 진단하고 구체적인 개선안을 제시해 드려요."
            ).style("font-size:11px; color: var(--dg-text-tertiary); margin-top:2px")

            with ui.column().classes("w-full gap-2 mt-2"):
                ui.label("소식글 제목 (현재 운영 중)").classes("dg-label-sm")
                title_input = ui.input(
                    placeholder="예: 유성궁동 변색렌즈 0원 한정 이벤트",
                ).props("outlined").classes("w-full dg-input")

                ui.label("소식글 본문 (현재 운영 중)").classes("dg-label-sm")
                newspost_input = ui.textarea(
                    placeholder=(
                        "현재 운영 중인 소식글 본문을 그대로 붙여넣으세요. "
                        "AI가 진단과 함께 완성형 수정안(제목·본문 전문)을 제시해 드려요."
                    ),
                ).props("rows=6 outlined").classes("w-full dg-input")

                ui.label("썸네일 / 광고 이미지 (선택)").classes("dg-label-sm")
                thumb_status = ui.label("(첨부 안 함)").style(
                    "font-size:12px; color: var(--dg-text-tertiary)"
                )
                ui.upload(
                    label="이미지 업로드 (.png/.jpg/.webp)",
                    auto_upload=True,
                    on_upload=lambda e: _handle_thumb_upload(e),
                    max_file_size=8_000_000,
                ).classes("max-w-md dg-upload").props('accept=".png,.jpg,.jpeg,.webp"')

            run_spinner = ui.row().classes("items-center gap-2 mt-2 hidden")
            with run_spinner:
                ui.spinner("dots", size="sm")
                run_step = ui.label("준비하고 있어요...").classes("dg-progress-text")

        # Step 3: Results
        results_card = ui.card().classes("dg-card w-full hidden")
        with results_card:
            section_header("summarize", "3단계: 분석 결과",
                           "현황 → 개선점 → 실행 계획 → 광고주 보고서")
            results_body = ui.column().classes("w-full gap-4")
            with ui.row().classes("gap-3 mt-3"):
                dl_default_btn = ui.button(
                    "보고서 DOCX 저장 (기본 폴더)", icon="save",
                    on_click=lambda: _export_docx(saveas=False),
                ).classes("dg-btn-success")
                dl_saveas_btn = ui.button(
                    "다른 위치로 저장...", icon="save_as",
                    on_click=lambda: _export_docx(saveas=True),
                ).classes("dg-btn-secondary")

        # ───────── handlers ─────────

        async def _handle_upload(e: Any) -> None:
            upload_status.clear()
            file_bytes = await e.file.read() if hasattr(e, "file") else e.content.read()
            filename = (getattr(e, "file", None) and e.file.name) or getattr(e, "name", "") or "uploaded.xlsx"
            state["filename"] = filename
            _log.info("Upload: %s (%d bytes)", filename, len(file_bytes))

            loop = asyncio.get_running_loop()
            try:
                parsed = await loop.run_in_executor(None, parse_demographic_xlsx, file_bytes)
            except Exception as exc:  # noqa: BLE001
                with upload_status:
                    with ui.element("div").classes("dg-banner dg-banner-error w-full"):
                        ui.icon("error", size="18px")
                        ui.label(f"파일을 읽지 못했습니다: {exc}")
                state["parsed"] = None
                run_btn.props("disabled")
                return

            n_age = len(parsed["ages"])
            n_camp = len(parsed["campaigns"])
            n_gender = len(parsed["genders"])

            if n_age == 0 and n_camp == 0 and n_gender == 0:
                sheets = _inspect_xlsx(file_bytes)
                with upload_status:
                    with ui.element("div").classes("dg-banner dg-banner-warning w-full"):
                        ui.icon("warning", size="18px")
                        with ui.column().classes("gap-1"):
                            ui.label(
                                "이 파일에서 연령/캠페인 데이터를 찾지 못했습니다."
                            ).style("font-weight:600")
                            ui.label(
                                "고급 분석에는 [연령 × 캠페인 breakdown] 양식이 필요해요. "
                                "날짜별 시계열 보고서는 [성과 보고서] 페이지에서 확인할 수 있어요."
                            ).style("font-size:12px")
                    if sheets:
                        with ui.card().classes("w-full"):
                            ui.label("업로드 파일에서 발견된 시트:").style("font-weight:600; font-size:13px")
                            for s in sheets:
                                h = ", ".join(s["headers"]) if s["headers"] else "(빈 시트)"
                                ui.label(f"• [{s['name']}]  →  {h}").style(
                                    "font-size:12px; color: var(--dg-text-tertiary)"
                                )
                state["parsed"] = None
                run_btn.props("disabled")
                return

            state["parsed"] = parsed
            with upload_status:
                with ui.element("div").classes("dg-banner dg-banner-success w-full"):
                    ui.icon("check_circle", size="18px")
                    ui.label(
                        f"파싱 완료 — 연령 {n_age}개 / 캠페인 {n_camp}개 / 성별 {n_gender}개. "
                        f"아래 [분석 실행] 버튼을 누르세요."
                    )
            run_btn.props(remove="disabled")

        async def _handle_thumb_upload(e: Any) -> None:
            try:
                data = await e.file.read() if hasattr(e, "file") else e.content.read()
                filename = (getattr(e, "file", None) and e.file.name) or getattr(e, "name", "") or "thumb"
                mime = getattr(getattr(e, "file", None), "content_type", None) or "image/png"
                state["thumbnail_data"] = data
                state["thumbnail_mime"] = mime
                state["thumbnail_name"] = filename
                thumb_status.set_text(f"첨부됨: {filename} ({len(data):,} bytes)")
                thumb_status.style("color: var(--dg-success); font-weight:600")
            except Exception as exc:  # noqa: BLE001
                ui.notify(f"이미지를 올리지 못했어요. 파일을 확인하고 다시 시도해 주세요. ({exc})", type="negative")

        async def _run_analysis() -> None:
            if not state["parsed"]:
                ui.notify("분석할 xlsx 파일을 먼저 올려 주세요.", type="warning")
                return

            parsed = state["parsed"]
            ages = parsed["ages"]
            campaigns = parsed["campaigns"]
            engine = engine_sel.value
            extra = extra_input.value or ""

            run_btn.props("disabled")
            run_spinner.classes(remove="hidden")
            run_step.set_text("1/3 규칙 기반으로 판정하고 있어요...")

            try:
                loop = asyncio.get_running_loop()
                judgments = judge_campaigns(campaigns)
                # Pass ages so simulation and priority include age-level OFF/boost
                plan = simulate_reallocation(judgments, ages=ages)
                priority = build_priority_checklist(judgments, ages=ages)
                var_warnings = check_variable_control(campaigns)
                pair_gaps = check_auto_manual_pairing(campaigns)

                # Funnel + economics
                funnel = calc_funnel(ages) if ages else calc_funnel(campaigns)
                aov = int(aov_input.value or 0)
                margin_pct = float(margin_input.value or 0)
                total_cost = sum(a.cost for a in ages) if ages else sum(c.cost for c in campaigns)
                total_actions = funnel.actions
                economics = calc_economics(
                    total_cost, total_actions,
                    avg_order_value=aov,
                    target_margin_rate=margin_pct / 100.0,
                )

                state["judgments"] = judgments
                state["plan"] = plan
                state["priority"] = priority
                state["var_warnings"] = var_warnings
                state["pair_gaps"] = pair_gaps
                state["funnel"] = funnel
                state["economics"] = economics

                pid = project_sel.value
                project = get_project(pid) if pid else {}
                project = project or {"name": "광고주", "campaign_name": state.get("filename", "")}

                engine_name = {"gpt": "GPT", "coordinate": "Claude+GPT 조율"}.get(engine, "Claude")
                run_step.set_text(f"2/3 {engine_name}가 보고서를 작성하고 있어요...")
                newspost_title = (title_input.value or "").strip()
                newspost_text = (newspost_input.value or "").strip()
                has_thumbnail = bool(state.get("thumbnail_data"))
                prompt = build_analysis_prompt(
                    project=project, ages=ages, campaigns=campaigns,
                    judgments=judgments, plan=plan, priority=priority,
                    var_warnings=var_warnings, pair_gaps=pair_gaps,
                    extra=extra, funnel=funnel, economics=economics,
                    newspost_title=newspost_title,
                    newspost_text=newspost_text,
                    has_thumbnail=has_thumbnail,
                )
                thumb_bytes = state.get("thumbnail_data")
                thumb_mime = state.get("thumbnail_mime", "image/png")

                if engine == "coordinate":
                    from app.ai.coordination import synthesize
                    claude_p = get_provider("claude")
                    gpt_p = get_provider("gpt")

                    def _claude_draft():
                        # 썸네일은 멀티모달 지원하는 Claude 초안에만 전달
                        if thumb_bytes:
                            try:
                                return claude_p.generate_text(
                                    prompt, system_prompt=SYSTEM_GUIDE_ANALYSIS,
                                    image=thumb_bytes, image_mime=thumb_mime,
                                )
                            except TypeError:
                                pass  # provider가 image 인자 미지원 → 텍스트 전용
                        return claude_p.generate_text(prompt, system_prompt=SYSTEM_GUIDE_ANALYSIS)

                    c_text, g_text = await asyncio.gather(
                        loop.run_in_executor(None, _claude_draft),
                        loop.run_in_executor(None, lambda: gpt_p.generate_text(prompt, system_prompt=SYSTEM_GUIDE_ANALYSIS)),
                    )
                    content = await loop.run_in_executor(
                        None, lambda: synthesize(c_text, g_text, "당근 광고 성과 분석"),
                    )
                else:
                    provider = get_provider(engine)
                    thumb_for_engine = thumb_bytes if engine == "claude" else None
                    if thumb_for_engine:
                        # Multi-modal call (Claude CLI stream-json)
                        content = await loop.run_in_executor(
                            None,
                            lambda: provider.generate_text(
                                prompt, system_prompt=SYSTEM_GUIDE_ANALYSIS,
                                image=thumb_for_engine, image_mime=thumb_mime,
                            ),
                        )
                    else:
                        content = await loop.run_in_executor(
                            None,
                            lambda: provider.generate_text(prompt, system_prompt=SYSTEM_GUIDE_ANALYSIS),
                        )
                sections = parse_analysis_sections(content)
                state["ai_sections"] = sections
                state["ai_raw"] = content

                run_step.set_text("3/3 결과 화면을 그리고 있어요...")
                _render_results(
                    ages=ages, campaigns=campaigns, judgments=judgments,
                    plan=plan, priority=priority,
                    var_warnings=var_warnings, pair_gaps=pair_gaps,
                    sections=sections, funnel=funnel, economics=economics,
                )
                results_card.classes(remove="hidden")
                ui.notify("분석이 끝났어요. 결과를 확인하고 보고서를 내려받아 보세요.", type="positive")
            except Exception as exc:  # noqa: BLE001
                _log.exception("Analysis failed")
                ui.notify(f"분석을 끝내지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative", timeout=8000)
            finally:
                run_spinner.classes("hidden")
                run_btn.props(remove="disabled")

        def _render_results(*, ages, campaigns, judgments, plan, priority,
                            var_warnings, pair_gaps, sections,
                            funnel=None, economics=None) -> None:
            results_body.clear()
            with results_body:

                # ── 한 줄 요약 ──
                summary = sections.get("summary", "").strip()
                if summary:
                    with ui.card().classes("w-full").style(
                        "background: var(--dg-primary-light); border-left:5px solid var(--dg-primary)"
                    ):
                        ui.label("핵심 요약").style(
                            "font-size:12px; font-weight:600; color: var(--dg-primary)"
                        )
                        ui.label(summary).style(
                            "font-size:16px; font-weight:700; color: var(--dg-text-primary); margin-top:4px"
                        )

                # ── 현황 진단 ──
                ui.label("현황 진단").classes("dg-section-title").style("margin-top:16px")
                _render_status_grid(ages, campaigns, plan)

                # 퍼널 + 경제성
                if funnel is not None and funnel.impressions > 0:
                    _render_funnel_card(funnel)
                if economics is not None and economics.avg_order_value > 0:
                    _render_economics_card(economics)

                if sections.get("status"):
                    with ui.card().classes("w-full"):
                        ui.markdown(sections["status"]).classes("dg-prose")

                # ── 개선점 ──
                ui.label("개선점").classes("dg-section-title").style("margin-top:16px")
                _render_action_cards(judgments, var_warnings, pair_gaps)
                if sections.get("findings"):
                    with ui.card().classes("w-full"):
                        ui.markdown(sections["findings"]).classes("dg-prose")

                # ── 캠페인 수정표 ──
                rev_rows = build_campaign_revision_table(judgments, plan) if judgments else []
                if rev_rows:
                    ui.label("캠페인 수정표 — 오늘 이대로 수정하면 돼요").classes(
                        "dg-section-title"
                    ).style("margin-top:16px")
                    _render_revision_table(rev_rows)

                # ── 실행 계획 ──
                ui.label("실행 계획").classes("dg-section-title").style("margin-top:16px")
                _render_priority_list(priority)
                if sections.get("plan"):
                    with ui.card().classes("w-full"):
                        ui.markdown(sections["plan"]).classes("dg-prose")

                # ── 예상 효과 ──
                ui.label("예상 효과").classes("dg-section-title").style("margin-top:16px")
                _render_realloc_summary(plan)
                if sections.get("expected"):
                    with ui.card().classes("w-full"):
                        ui.markdown(sections["expected"]).classes("dg-prose")

                # ── 광고주 전달 멘트 ──
                if sections.get("client_note"):
                    ui.label("광고주 전달 멘트").classes("dg-section-title").style("margin-top:16px")
                    with ui.card().classes("w-full").style(
                        "background: #FFF8E1; border-left:5px solid var(--dg-warning)"
                    ):
                        ui.markdown(sections["client_note"]).classes("dg-prose")

                # ── 소식글 카피 수정안 ──
                if sections.get("copy_revisions"):
                    ui.label("소식글 카피 수정안 (제목 + 본문 완성형)").classes(
                        "dg-section-title"
                    ).style("margin-top:16px")
                    with ui.card().classes("w-full").style(
                        "border-left:5px solid var(--dg-primary)"
                    ):
                        ui.markdown(sections["copy_revisions"]).classes("dg-prose")
                        ui.label(
                            "위 본문을 그대로 당근 비즈프로필 소식글에 복붙해 운영해보세요."
                        ).style("font-size:11px; color: var(--dg-text-tertiary); margin-top:8px")

                # ── 썸네일 이미지 생성 프롬프트 ──
                if sections.get("thumbnail_prompts"):
                    ui.label("썸네일 이미지 생성 프롬프트 (Gemini/Nano Banana 복붙용)").classes(
                        "dg-section-title"
                    ).style("margin-top:16px")
                    with ui.card().classes("w-full").style(
                        "border-left:5px solid var(--dg-success)"
                    ):
                        ui.markdown(sections["thumbnail_prompts"]).classes("dg-prose")
                        ui.label(
                            "각 프롬프트 블록을 복사해 이미지 생성기에 그대로 붙여넣으세요. "
                            "이미지는 썸네일 페이지(/thumbnail)나 외부 도구에서 만들 수 있어요."
                        ).style("font-size:11px; color: var(--dg-text-tertiary); margin-top:8px")

                # ── 연령 묶음 추천 ──
                if ages:
                    ui.label("연령 묶음 추천 (캠페인 분리 가이드)").classes(
                        "dg-section-title"
                    ).style("margin-top:16px")
                    _render_age_groups(ages)

                # ── 원본 데이터 ──
                with ui.expansion("원본 데이터 표 보기", icon="table_chart").classes("w-full"):
                    if ages:
                        ui.label("연령대 breakdown").style("font-weight:600; margin-top:8px")
                        ui.table(
                            columns=[
                                {"name": "연령", "label": "연령", "field": "연령", "align": "left"},
                                {"name": "비용", "label": "비용", "field": "비용", "align": "right"},
                                {"name": "행동", "label": "행동", "field": "행동", "align": "right"},
                                {"name": "CPA", "label": "CPA", "field": "CPA", "align": "right"},
                                {"name": "CTR", "label": "CTR", "field": "CTR", "align": "right"},
                            ],
                            rows=[
                                {"연령": a.label, "비용": _fmt_won(a.cost),
                                 "행동": a.actions,
                                 "CPA": _fmt_won(a.cpa) if a.actions else "-",
                                 "CTR": f"{a.ctr:.2f}%"}
                                for a in ages
                            ],
                        ).classes("w-full dg-table").props("dense flat bordered")
                    if judgments:
                        ui.label("캠페인 판정").style("font-weight:600; margin-top:8px")
                        ui.table(
                            columns=[
                                {"name": "캠페인", "label": "캠페인", "field": "캠페인", "align": "left"},
                                {"name": "입찰", "label": "입찰", "field": "입찰", "align": "center"},
                                {"name": "비용", "label": "비용", "field": "비용", "align": "right"},
                                {"name": "행동", "label": "행동", "field": "행동", "align": "right"},
                                {"name": "CPA", "label": "CPA", "field": "CPA", "align": "right"},
                                {"name": "판정", "label": "판정", "field": "판정", "align": "center"},
                                {"name": "사유", "label": "사유", "field": "사유", "align": "left"},
                            ],
                            rows=[
                                {"캠페인": j.campaign.name,
                                 "입찰": {"manual": "수동", "auto": "자동"}.get(j.campaign.bid_mode, "-"),
                                 "비용": _fmt_won(j.campaign.cost), "행동": j.campaign.actions,
                                 "CPA": _fmt_won(j.campaign.cpa) if j.campaign.actions else "-",
                                 "판정": j.verdict, "사유": j.reason}
                                for j in judgments
                            ],
                        ).classes("w-full dg-table").props("dense flat bordered")

        def _render_funnel_card(funnel) -> None:
            stages = [
                ("노출", funnel.impressions, 100.0, "#7295C4"),
                ("클릭", funnel.clicks, funnel.ctr, "#E08F55"),
                ("행동", funnel.actions, funnel.cvr, "#6FAE8F"),
            ]
            with ui.card().classes("w-full"):
                ui.label("퍼널 단계별 전환").style(
                    "font-size:14px; font-weight:600; color: var(--dg-text-primary)"
                )
                with ui.row().classes("w-full items-stretch gap-2 mt-2"):
                    for i, (label, count, rate, color) in enumerate(stages):
                        if i > 0:
                            ui.label("→").style(
                                "font-size:20px; align-self:center; color: var(--dg-text-tertiary)"
                            )
                        with ui.element("div").style(
                            f"flex:1; padding:10px; border-radius:8px; "
                            f"background:{color}11; border:1px solid {color}55; "
                            f"text-align:center"
                        ):
                            ui.label(f"{count:,}").style(
                                f"font-size:20px; font-weight:700; color:{color}"
                            )
                            if i == 0:
                                ui.label("100%").style(
                                    f"font-size:11px; color:{color}; font-weight:600"
                                )
                            else:
                                ui.label(f"{rate:.2f}%").style(
                                    f"font-size:11px; color:{color}; font-weight:600"
                                )
                            ui.label(label).style(
                                "font-size:12px; color: var(--dg-text-secondary); margin-top:2px"
                            )

                if funnel.bottleneck:
                    drop = (funnel.drop_impression_to_click
                            if funnel.bottleneck == "노출→클릭"
                            else funnel.drop_click_to_action)
                    with ui.element("div").classes("dg-banner dg-banner-warning w-full mt-2"):
                        ui.icon("warning", size="16px")
                        ui.label(
                            f"최대 이탈 구간: {funnel.bottleneck} ({drop:.1f}% 이탈) — "
                            f"이 구간 개선이 가장 큰 효과를 냅니다."
                        )

        def _render_economics_card(economics) -> None:
            status_color = {
                "profit": "var(--dg-success)",
                "breakeven": "var(--dg-warning)",
                "loss": "var(--dg-error)",
                "unknown": "var(--dg-text-tertiary)",
            }[economics.status]
            status_label = {
                "profit": "흑자 / 확장 여력",
                "breakeven": "손익분기 근접",
                "loss": "적자 / 즉시 조정 필요",
                "unknown": "판단 보류",
            }[economics.status]

            with ui.card().classes("w-full").style(f"border-left:5px solid {status_color}"):
                ui.label("MAX CPA · 손익 분석").style(
                    "font-size:14px; font-weight:600; color: var(--dg-text-primary)"
                )
                ui.label(
                    f"객단가 {economics.avg_order_value:,}원 · "
                    f"목표 이익률 {economics.target_margin_rate*100:.0f}% 기준"
                ).style("font-size:11px; color: var(--dg-text-tertiary); margin-top:2px")

                with ui.row().classes("w-full gap-6 flex-wrap mt-2"):
                    for label, val, accent in [
                        ("현재 CPA", _fmt_won(economics.current_cpa), False),
                        ("손익분기 CPA", _fmt_won(economics.breakeven_cpa), False),
                        ("MAX CPA (목표 이익 반영)", _fmt_won(economics.max_cpa), True),
                        ("한계 소진율", f"{economics.burn_rate*100:.1f}%", True),
                        ("판정", status_label, True),
                    ]:
                        with ui.column().classes("gap-0"):
                            ui.label(label).style(
                                "font-size:11px; color: var(--dg-text-tertiary)"
                            )
                            ui.label(val).style(
                                f"font-size:16px; font-weight:700; "
                                f"color:{status_color if accent else 'var(--dg-text-primary)'}"
                            )

                ui.label(
                    f"예상 매출 {economics.expected_revenue:,}원 / "
                    f"광고 후 이익 {economics.expected_profit:,}원"
                ).style("font-size:12px; color: var(--dg-text-secondary); margin-top:6px")

        def _render_status_grid(ages, campaigns, plan) -> None:
            total_cost = sum(a.cost for a in ages) or sum(c.cost for c in campaigns)
            total_actions = sum(a.actions for a in ages) or sum(c.actions for c in campaigns)
            avg_cpa = total_cost / total_actions if total_actions else 0
            active_ages = [a for a in ages if a.actions > 0]
            best = min(active_ages, key=lambda a: a.cpa) if active_ages else None
            worst = max(active_ages, key=lambda a: a.cpa) if active_ages else None

            kpis = [
                ("총 비용", _fmt_won(total_cost), False),
                ("총 행동", f"{total_actions:,}건", False),
                ("평균 CPA", _fmt_won(avg_cpa) if total_actions else "-", True),
                ("캠페인 수", f"{len(campaigns)}개", False),
            ]
            if best:
                kpis.append(("최고 효율 연령", f"{best.label} ({_fmt_won(best.cpa)})", True))
            if worst and worst.label != (best.label if best else ""):
                kpis.append(("최악 효율 연령", f"{worst.label} ({_fmt_won(worst.cpa)})", True))

            with ui.element("div").classes("dg-kpi-grid w-full").style(
                "display:grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap:12px"
            ):
                for label, val, accent in kpis:
                    with ui.element("div").classes("dg-kpi-card"):
                        ui.label(label).classes("dg-kpi-label")
                        ui.label(val).classes("dg-kpi-value-accent" if accent else "dg-kpi-value")

        def _render_revision_table(rev_rows) -> None:
            """판정+재배분을 합친 실행용 수정표. 우선순위 순서대로 따라 하면 된다."""
            with ui.card().classes("w-full").style("border-left: 5px solid var(--dg-primary)"):
                ui.table(
                    columns=REVISION_TABLE_COLUMNS,
                    rows=revision_rows_for_table(rev_rows),
                    row_key="target",
                ).classes("w-full dg-table").props("dense flat bordered wrap-cells")
                with ui.row().classes("gap-3 mt-2 items-center"):
                    def _copy_revision() -> None:
                        ui.clipboard.write(revision_table_markdown(rev_rows))
                        ui.notify(
                            "수정표를 마크다운으로 복사했어요. 보고서나 메모에 붙여 넣어 보세요.",
                            type="positive", timeout=2500,
                        )
                    ui.button(
                        "마크다운 복사", icon="content_copy",
                        on_click=_copy_revision,
                    ).classes("dg-btn-secondary dg-btn-sm")
                    ui.label("우선순위 1부터 순서대로 처리하면 돼요.").classes("dg-label-sm")

        def _render_action_cards(judgments, var_warnings, pair_gaps) -> None:
            actionable = [j for j in judgments
                          if j.verdict in ("캠페인OFF", "소재전면교체", "증액")]
            if not actionable:
                with ui.card().classes("w-full").style(
                    "background: var(--dg-success-light); border-left:4px solid var(--dg-success)"
                ):
                    ui.icon("check_circle", size="20px").style("color: var(--dg-success)")
                    ui.label(
                        "캠페인 단위로는 지금 바로 조치할 항목이 없어요. "
                        "연령 단위 액션(우선순위 체크리스트)과 AI 분석 결과를 참고하세요."
                    ).style("font-weight:600; font-size:13px")
                return

            with ui.element("div").style(
                "display:grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap:12px"
            ):
                for j in actionable:
                    color, label = _VERDICT_BADGE.get(j.verdict, ("info", j.verdict))
                    with ui.card().classes("w-full").style(
                        f"border-left:5px solid var(--dg-{color})"
                    ):
                        ui.label(label).style(
                            f"font-size:11px; font-weight:700; color: var(--dg-{color}); "
                            f"text-transform:uppercase; letter-spacing:0.5px"
                        )
                        ui.label(j.campaign.name).style(
                            "font-size:14px; font-weight:600; margin-top:2px"
                        )
                        ui.label(j.reason).style(
                            "font-size:12px; color: var(--dg-text-secondary); line-height:1.5; margin-top:4px"
                        )

                # 캠페인 비교 경고는 표시하지 않습니다 (사용자 피드백: 비교가 아닌 개별 분석).

        def _render_priority_list(priority) -> None:
            if not priority:
                return
            with ui.card().classes("w-full"):
                for i, item in enumerate(priority):
                    with ui.row().classes("items-start gap-2 py-1"):
                        with ui.element("div").style(
                            "min-width:28px; height:28px; border-radius:14px; "
                            "background: var(--dg-primary); color:white; display:flex; "
                            "align-items:center; justify-content:center; font-weight:700; font-size:13px; "
                            "flex-shrink:0"
                        ):
                            ui.label(str(i + 1)).style("color:white; font-weight:700; font-size:13px")
                        ui.label(item).style(
                            "font-size:13px; line-height:1.6; padding-top:4px"
                        )

        def _render_realloc_summary(plan) -> None:
            with ui.card().classes("w-full").style("background: var(--dg-primary-light)"):
                with ui.row().classes("w-full gap-6 flex-wrap"):
                    for label, val, color in [
                        ("현재 총예산", _fmt_won(plan.current_total), "var(--dg-text-primary)"),
                        ("예상 절감액", _fmt_won(plan.savings), "var(--dg-success)"),
                        ("추가 행동 예상", f"+{plan.expected_action_delta}건", "var(--dg-primary)"),
                        ("조정 후 총예산", _fmt_won(plan.projected_total), "var(--dg-text-primary)"),
                    ]:
                        with ui.column().classes("gap-0"):
                            ui.label(label).style("font-size:11px; color: var(--dg-text-tertiary)")
                            ui.label(val).style(f"font-size:18px; font-weight:700; color:{color}")

                if plan.cuts:
                    ui.separator().style("margin: 8px 0")
                    ui.label("축소 / OFF 대상").style("font-size:12px; font-weight:600; color: var(--dg-error)")
                    for n, a in plan.cuts:
                        ui.label(f"  • {n}: {_fmt_won(a)} 절감").style("font-size:12px")
                if plan.boosts:
                    ui.label("증액 대상").style("font-size:12px; font-weight:600; color: var(--dg-success)")
                    for n, a in plan.boosts:
                        ui.label(f"  • {n}: +{_fmt_won(a)}").style("font-size:12px")

        def _render_age_groups(ages) -> None:
            group_count = {"value": 3}

            def _refresh() -> None:
                groups_container.clear()
                groups = group_ages_by_cpa(ages, n_groups=group_count["value"])
                with groups_container:
                    ui.label(
                        f"이 {len(groups)}개 묶음을 각각 별도 캠페인으로 분리하고, "
                        f"수동+자동 페어로 운영하세요."
                    ).style("font-size:12px; color: var(--dg-text-secondary); margin-bottom:6px")
                    for i, g in enumerate(groups):
                        color = ("var(--dg-success)" if i == 0
                                 else "var(--dg-primary)" if g.avg_cpa != float("inf")
                                 else "var(--dg-error)")
                        cpa_str = "OFF 권장" if g.avg_cpa == float("inf") else _fmt_won(g.avg_cpa)
                        with ui.card().style(f"border-left:4px solid {color}; width:100%"):
                            with ui.row().classes("w-full items-center justify-between"):
                                ui.label(f"묶음 {i+1} — {' / '.join(g.members)}").style(
                                    "font-weight:600; font-size:13px"
                                )
                                ui.label(f"평균 CPA {cpa_str}").style(
                                    f"color:{color}; font-weight:700; font-size:13px"
                                )

            with ui.row().classes("w-full items-center gap-3"):
                ui.label("묶음 수").style("font-size:13px")
                slider = ui.slider(min=2, max=5, value=3, step=1).classes("w-48").props("label-always")
                slider.bind_value(group_count, "value")
                slider.on_value_change(lambda _e: _refresh())
            groups_container = ui.column().classes("w-full gap-2")
            _refresh()

        async def _export_docx(*, saveas: bool) -> None:
            if not state.get("ai_sections"):
                ui.notify("[분석 실행]을 먼저 눌러 주세요.", type="warning")
                return
            try:
                pid = project_sel.value
                project = get_project(pid) if pid else {}
                project = project or {}
                project_meta = _build_project_meta(project, state)
                economics = state.get("economics")
                funnel = state.get("funnel")
                parsed = state["parsed"]
                loop = asyncio.get_running_loop()
                import tempfile
                with tempfile.TemporaryDirectory() as tmpdir:
                    tmp = Path(tmpdir)
                    out = await loop.run_in_executor(
                        None,
                        lambda: build_analysis_docx(
                            project_meta=project_meta,
                            ages=parsed["ages"], campaigns=parsed["campaigns"],
                            judgments=state["judgments"], plan=state["plan"],
                            priority=state["priority"],
                            var_warnings=state["var_warnings"],
                            pair_gaps=state["pair_gaps"],
                            ai_sections=state["ai_sections"],
                            output_path=tmp / "analysis.docx",
                            chart_dir=CHARTS_DIR,
                            funnel=funnel,
                            economics=economics,
                        ),
                    )
                    data = out.read_bytes()

                fname = f"고급분석_{project_meta['name']}.docx".replace("/", "_")
                if saveas:
                    ok = await ExportManager.save_as_multi([(data, fname)])
                    if ok:
                        ui.notify("저장했어요.", type="positive")
                else:
                    ExportManager.save_default(data, filename=fname)
                    ui.notify(f"{fname} 다운로드를 시작했어요.", type="positive", timeout=6000)
            except Exception as exc:  # noqa: BLE001
                _log.exception("DOCX export failed")
                ui.notify(f"보고서를 만들지 못했어요. 잠시 후 다시 시도해 주세요. ({exc})", type="negative", timeout=8000)
