"""Demographic analysis for Daangn ads.

Input contract: demographic breakdown rows from 당근 광고 관리자 내보내기.
Output: insights (최고효율/숨겨진기회/예산불균형), age groupings,
campaign judgments (유지/OFF/교체/증액), budget reallocation simulation,
variable control warnings, and auto/manual pairing checks.

This module contains pure logic (no I/O, no UI) so it can be unit-tested
independently of NiceGUI.

Based on operator playbook:
- 당근은 머신러닝 체감상 없음 → 연령/성별 직접 찢기
- 수동+자동 캠페인 동시 운영 (수동=성과, 자동=노출부스팅)
- 변수 통제 원칙: 캠페인 간 1개만 다르게
- 연령 그룹핑은 행동당비용 유사도 기준
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Literal, Sequence


# ───────────────────────── dataclasses ─────────────────────────


@dataclass(frozen=True)
class Segment:
    """One row of demographic breakdown (성별 또는 연령대)."""

    label: str  # "남성", "여성", "15-19", "20-24", ...
    cost: int
    actions: int  # 총행동 (문의+단골+쿠폰 합)
    impressions: int = 0
    clicks: int = 0

    @property
    def cpa(self) -> float:
        return self.cost / self.actions if self.actions > 0 else 0.0

    @property
    def ctr(self) -> float:
        return (self.clicks / self.impressions * 100) if self.impressions > 0 else 0.0


@dataclass(frozen=True)
class CampaignPerf:
    """One campaign's performance row."""

    name: str
    cost: int
    actions: int
    creative_count: int = 1
    impressions: int = 0
    clicks: int = 0
    # Optional metadata for variable control + auto/manual pairing:
    bid_mode: Literal["auto", "manual", "unknown"] = "unknown"
    age_range: str = ""  # "20-29" etc.
    creative_id: str = ""

    @property
    def cpa(self) -> float:
        return self.cost / self.actions if self.actions > 0 else 0.0

    @property
    def ctr(self) -> float:
        return (self.clicks / self.impressions * 100) if self.impressions > 0 else 0.0

    @property
    def cost_share(self) -> float:
        # Filled in by analyze_campaigns() since it requires total
        return 0.0


@dataclass(frozen=True)
class Insight:
    kind: Literal["best_efficiency", "hidden_opportunity", "budget_imbalance"]
    label: str
    message: str
    numbers: dict[str, float] = field(default_factory=dict)


@dataclass(frozen=True)
class AgeGroup:
    """A cluster of age segments with similar CPA."""

    members: tuple[str, ...]  # e.g. ("40-44", "55-59")
    avg_cpa: float
    total_cost: int
    total_actions: int


@dataclass(frozen=True)
class CampaignJudgment:
    campaign: CampaignPerf
    verdict: Literal["유지", "소재정리후유지", "소재전면교체", "캠페인OFF", "증액"]
    reason: str
    cost_share: float  # percent 0~100


@dataclass(frozen=True)
class ReallocationPlan:
    current_total: int
    projected_total: int  # after OFF/cuts
    savings: int
    expected_action_delta: int  # estimated +N inquiries from reallocation
    cuts: tuple[tuple[str, int], ...]  # (campaign_name, cut_amount)
    boosts: tuple[tuple[str, int], ...]  # (campaign_name, boost_amount)


@dataclass(frozen=True)
class VariableControlWarning:
    campaign_a: str
    campaign_b: str
    diff_count: int
    diffs: tuple[str, ...]  # names of differing attributes


@dataclass(frozen=True)
class PairingGap:
    campaign: str
    missing_counterpart: Literal["auto", "manual"]


# ───────────────────────── demographic analysis ─────────────────────────


def analyze_segments(segments: Sequence[Segment]) -> list[Insight]:
    """Produce insights for a set of demographic segments (either 성별 or 연령).

    Rules:
    - best_efficiency: lowest CPA AND CTR ≥ average
    - hidden_opportunity: CPA below median but cost share < 5%
    - budget_imbalance: cost share > average AND CPA above median
    """
    active = [s for s in segments if s.actions > 0 and s.cost > 0]
    if not active:
        return []

    total_cost = sum(s.cost for s in active)
    if total_cost <= 0:
        return []

    cpas = sorted(s.cpa for s in active)
    median_cpa = cpas[len(cpas) // 2]
    avg_ctr = sum(s.ctr for s in active) / len(active) if active else 0.0

    insights: list[Insight] = []

    best = min(active, key=lambda s: s.cpa)
    if best.ctr >= avg_ctr:
        insights.append(
            Insight(
                kind="best_efficiency",
                label=best.label,
                message=f"최고 효율 {best.label} (행동당비용 {int(best.cpa):,}원, CTR {best.ctr:.2f}%)",
                numbers={"cpa": best.cpa, "ctr": best.ctr},
            )
        )

    for s in active:
        share_pct = s.cost / total_cost * 100
        if s.cpa <= median_cpa and share_pct < 5.0:
            insights.append(
                Insight(
                    kind="hidden_opportunity",
                    label=s.label,
                    message=f"숨겨진 기회: {s.label} (행동당비용 {int(s.cpa):,}원이나 비용비중 {share_pct:.1f}%로 낮음)",
                    numbers={"cpa": s.cpa, "cost_share": share_pct},
                )
            )

    avg_share = 100.0 / len(active)
    for s in active:
        share_pct = s.cost / total_cost * 100
        if share_pct > avg_share and s.cpa > median_cpa:
            insights.append(
                Insight(
                    kind="budget_imbalance",
                    label=s.label,
                    message=f"예산 불균형: {s.label} (비용비중 {share_pct:.1f}%인데 행동당비용 {int(s.cpa):,}원으로 비효율)",
                    numbers={"cpa": s.cpa, "cost_share": share_pct},
                )
            )

    return insights


# ───────────────────────── age grouping ─────────────────────────


def group_ages_by_cpa(
    age_segments: Sequence[Segment],
    n_groups: int = 3,
) -> list[AgeGroup]:
    """Cluster age segments by CPA similarity using 1D k-means.

    Returns groups sorted by avg_cpa ascending (가장 효율 좋은 묶음이 첫번째).
    n_groups must be 2~5. Segments with 0 actions are placed in a catch-all
    group at the end (cpa=inf, recommend OFF).
    """
    if n_groups < 2:
        n_groups = 2
    if n_groups > 5:
        n_groups = 5

    active = [s for s in age_segments if s.actions > 0]
    inactive = [s for s in age_segments if s.actions == 0 and s.cost > 0]

    groups: list[AgeGroup] = []

    if active:
        n = min(n_groups, len(active))
        clusters = _kmeans_1d([s.cpa for s in active], n)

        for cluster_indices in clusters:
            members = tuple(active[i].label for i in cluster_indices)
            total_cost = sum(active[i].cost for i in cluster_indices)
            total_actions = sum(active[i].actions for i in cluster_indices)
            avg_cpa = total_cost / total_actions if total_actions > 0 else 0.0
            groups.append(
                AgeGroup(
                    members=members,
                    avg_cpa=avg_cpa,
                    total_cost=total_cost,
                    total_actions=total_actions,
                )
            )

        groups.sort(key=lambda g: g.avg_cpa)

    if inactive:
        members = tuple(s.label for s in inactive)
        total_cost = sum(s.cost for s in inactive)
        groups.append(
            AgeGroup(
                members=members,
                avg_cpa=float("inf"),
                total_cost=total_cost,
                total_actions=0,
            )
        )

    return groups


def _kmeans_1d(values: list[float], k: int, max_iter: int = 50) -> list[list[int]]:
    """Simple 1D k-means. Returns list of index-lists (clusters)."""
    if not values or k <= 0:
        return []

    k = min(k, len(values))
    sorted_idx = sorted(range(len(values)), key=lambda i: values[i])

    # Initialize centroids evenly across sorted values
    step = max(1, len(values) // k)
    centroids = [values[sorted_idx[min(i * step, len(values) - 1)]] for i in range(k)]

    assignments = [0] * len(values)
    for _ in range(max_iter):
        changed = False
        for i, v in enumerate(values):
            best = min(range(k), key=lambda c: abs(v - centroids[c]))
            if assignments[i] != best:
                assignments[i] = best
                changed = True

        new_centroids = []
        for c in range(k):
            members = [values[i] for i in range(len(values)) if assignments[i] == c]
            new_centroids.append(sum(members) / len(members) if members else centroids[c])

        if not changed:
            break
        centroids = new_centroids

    clusters: list[list[int]] = [[] for _ in range(k)]
    for i, a in enumerate(assignments):
        clusters[a].append(i)
    return [c for c in clusters if c]


# ───────────────────────── campaign judgment ─────────────────────────


def judge_campaigns(
    campaigns: Sequence[CampaignPerf],
    *,
    min_actions: int = 10,
    cpa_replace_multiplier: float = 1.5,
    cost_share_high: float = 20.0,
    cost_share_low: float = 5.0,
    boost_cpa_multiplier: float = 0.6,
) -> list[CampaignJudgment]:
    """Judge each campaign using action count × CPA × cost share.

    Priority order (first match wins):
    1. actions < min_actions + spent significant → 캠페인OFF (dead campaign)
    2. CPA > avg_cpa × cpa_replace_multiplier → 소재전면교체 (creative problem)
    3. share ≥ cost_share_high AND CPA ≤ avg_cpa → 소재정리후유지 (주력 유지)
    4. CPA < avg_cpa × boost_cpa_multiplier AND share < cost_share_low → 증액 (hidden gem)
    5. else → 유지 (normal range)
    """
    active = [c for c in campaigns if c.cost > 0]
    if not active:
        return []

    total_cost = sum(c.cost for c in active)
    total_actions = sum(c.actions for c in active)
    avg_cpa = total_cost / total_actions if total_actions > 0 else 0.0

    judgments: list[CampaignJudgment] = []
    for c in active:
        share = c.cost / total_cost * 100
        verdict: str
        reason: str

        if c.actions == 0:
            verdict = "캠페인OFF"
            reason = f"행동 0건 ({c.cost:,}원 소진, 즉시 중단 권장)"
        elif c.actions < min_actions:
            verdict = "캠페인OFF"
            reason = f"행동 {c.actions}건 (기준 {min_actions}건 미만, CPA {int(c.cpa):,}원 비효율)"
        elif avg_cpa > 0 and c.cpa > avg_cpa * cpa_replace_multiplier:
            verdict = "소재전면교체"
            reason = f"CPA {int(c.cpa):,}원 (평균 {int(avg_cpa):,}원 대비 {c.cpa/avg_cpa:.1f}배 높음)"
        elif share >= cost_share_high and (avg_cpa == 0 or c.cpa <= avg_cpa):
            verdict = "소재정리후유지"
            reason = f"비중 {share:.1f}% + CPA {int(c.cpa):,}원 (평균 이하, 주력 유지)"
        elif avg_cpa > 0 and c.cpa < avg_cpa * boost_cpa_multiplier and share < cost_share_low:
            verdict = "증액"
            reason = f"CPA {int(c.cpa):,}원 (평균 {int(avg_cpa):,}원 대비 우수)인데 비중 {share:.1f}%로 작음"
        else:
            verdict = "유지"
            reason = f"평균 범위 (CPA {int(c.cpa):,}원, 비중 {share:.1f}%)"

        judgments.append(
            CampaignJudgment(
                campaign=c,
                verdict=verdict,  # type: ignore[arg-type]
                reason=reason,
                cost_share=share,
            )
        )

    return judgments


# ───────────────────────── budget reallocation ─────────────────────────


def simulate_reallocation(
    judgments: Sequence[CampaignJudgment],
    ages: Sequence[Segment] | None = None,
    *,
    age_off_cpa_multiplier: float = 2.0,
    age_boost_cpa_multiplier: float = 0.7,
) -> ReallocationPlan:
    """Build a reallocation plan based on judgments and optional age breakdown.

    Campaign-level rules:
      OFF: remove entire cost from budget.
      소재전면교체: cut 50% while testing new creatives.
      증액: double the budget (capped by total saved).

    Age-level rules (when `ages` provided — entries prefixed with "[연령]"):
      cost > 0 and actions == 0 → OFF (full age cost)
      cpa > avg × age_off_cpa_multiplier and actions < avg actions → OFF
      cpa < avg × age_boost_cpa_multiplier and actions > avg actions → 증액

    Expected action delta is estimated with each boosted unit's current CPA.
    """
    current_total = sum(j.campaign.cost for j in judgments)

    cuts: list[tuple[str, int]] = []
    boost_candidates: list[tuple[str, float, int]] = []  # (label, cpa, desired_boost)

    for j in judgments:
        if j.verdict == "캠페인OFF":
            cuts.append((j.campaign.name, j.campaign.cost))
        elif j.verdict == "소재전면교체":
            cuts.append((j.campaign.name, j.campaign.cost // 2))
        elif j.verdict == "증액":
            boost_candidates.append((j.campaign.name, j.campaign.cpa, j.campaign.cost))

    if ages:
        active = [a for a in ages if a.actions > 0]
        avg_cpa = (
            sum(a.cost for a in active) / sum(a.actions for a in active)
            if active and sum(a.actions for a in active) > 0
            else 0.0
        )
        avg_actions = sum(a.actions for a in active) / len(active) if active else 0.0
        for a in ages:
            if a.cost <= 0:
                continue
            if a.actions == 0:
                cuts.append((f"[연령] {a.label}", a.cost))
            elif (
                avg_cpa > 0
                and a.cpa > avg_cpa * age_off_cpa_multiplier
                and a.actions < avg_actions
            ):
                cuts.append((f"[연령] {a.label}", a.cost))
            elif (
                avg_cpa > 0
                and a.cpa < avg_cpa * age_boost_cpa_multiplier
                and a.actions > avg_actions
            ):
                boost_candidates.append((f"[연령] {a.label}", a.cpa, a.cost))

    total_savings = sum(amt for _, amt in cuts)
    total_boost_desired = sum(d for _, _, d in boost_candidates)

    boosts: list[tuple[str, int]] = []
    expected_delta = 0

    if total_boost_desired > 0 and total_savings > 0:
        ratio = min(1.0, total_savings / total_boost_desired)
        for name, cpa, desired in boost_candidates:
            applied = int(desired * ratio)
            boosts.append((name, applied))
            if cpa > 0:
                expected_delta += int(applied / cpa)

    projected_total = current_total - total_savings + sum(amt for _, amt in boosts)

    return ReallocationPlan(
        current_total=current_total,
        projected_total=projected_total,
        savings=total_savings - sum(amt for _, amt in boosts),
        expected_action_delta=expected_delta,
        cuts=tuple(cuts),
        boosts=tuple(boosts),
    )


# ───────────────────────── execution priority ─────────────────────────


def build_priority_checklist(
    judgments: Sequence[CampaignJudgment],
    ages: Sequence[Segment] | None = None,
    *,
    age_off_cpa_multiplier: float = 2.0,
    age_boost_cpa_multiplier: float = 0.7,
) -> list[str]:
    """Deterministic execution order.

    1순위: 비효율(캠페인/연령) 축소/OFF
    2순위: 고효율(캠페인/연령) 증액
    3순위: 소재 정리 및 A/B 테스트
    4순위: 신규 소재/타겟 테스트
    """
    off = [j.campaign.name for j in judgments if j.verdict == "캠페인OFF"]
    boost = [j.campaign.name for j in judgments if j.verdict == "증액"]
    replace = [j.campaign.name for j in judgments if j.verdict == "소재전면교체"]
    keep = [j.campaign.name for j in judgments if j.verdict == "소재정리후유지"]

    age_off: list[str] = []
    age_boost: list[str] = []
    if ages:
        active = [a for a in ages if a.actions > 0]
        avg_cpa = (
            sum(a.cost for a in active) / sum(a.actions for a in active)
            if active and sum(a.actions for a in active) > 0
            else 0.0
        )
        avg_actions = sum(a.actions for a in active) / len(active) if active else 0.0
        for a in ages:
            if a.cost <= 0:
                continue
            if a.actions == 0:
                age_off.append(a.label)
            elif (
                avg_cpa > 0
                and a.cpa > avg_cpa * age_off_cpa_multiplier
                and a.actions < avg_actions
            ):
                age_off.append(a.label)
            elif (
                avg_cpa > 0
                and a.cpa < avg_cpa * age_boost_cpa_multiplier
                and a.actions > avg_actions
            ):
                age_boost.append(a.label)

    items: list[str] = []
    off_labels = [f"캠페인 {n}" for n in off] + [f"연령 {n}" for n in age_off]
    if off_labels:
        items.append(f"1순위 — 비효율 축소/OFF: {', '.join(off_labels)}")
    boost_labels = [f"캠페인 {n}" for n in boost] + [f"연령 {n}" for n in age_boost]
    if boost_labels:
        items.append(f"2순위 — 고효율 예산 확대: {', '.join(boost_labels)}")
    if replace or keep:
        targets = replace + keep
        items.append(f"3순위 — 소재 정리 및 A/B 테스트: {', '.join(targets)}")
    items.append("4순위 — 신규 소재/타겟 테스트 (연령 그룹별 신규 캠페인 투입)")
    return items


# ───────────────────────── variable control ─────────────────────────


def check_variable_control(
    campaigns: Sequence[CampaignPerf],
) -> list[VariableControlWarning]:
    """Detect campaigns that differ by >1 variable (bid_mode/age_range/creative_id).

    Operator rule: 비교 목적으로 운영되는 캠페인 쌍은 1개 변수만 달라야 한다.
    Compares pairs sharing at least one attribute (grouping key) and flags the
    rest that differ.
    """
    warnings: list[VariableControlWarning] = []
    campaigns = [c for c in campaigns if c.cost > 0]

    for i in range(len(campaigns)):
        for j in range(i + 1, len(campaigns)):
            a, b = campaigns[i], campaigns[j]
            diffs: list[str] = []
            if a.bid_mode != b.bid_mode and "unknown" not in (a.bid_mode, b.bid_mode):
                diffs.append("bid_mode(자동/수동)")
            if a.age_range != b.age_range and a.age_range and b.age_range:
                diffs.append("age_range(연령)")
            if a.creative_id != b.creative_id and a.creative_id and b.creative_id:
                diffs.append("creative_id(소재)")

            if len(diffs) >= 2:
                warnings.append(
                    VariableControlWarning(
                        campaign_a=a.name,
                        campaign_b=b.name,
                        diff_count=len(diffs),
                        diffs=tuple(diffs),
                    )
                )
    return warnings


# ───────────────────────── auto/manual pairing ─────────────────────────


def check_auto_manual_pairing(
    campaigns: Sequence[CampaignPerf],
) -> list[PairingGap]:
    """Verify each 수동 campaign has a 자동 twin (same age_range + creative_id).

    Operator rule: 수동+자동을 같은 조건으로 동시 운영해야 노출이 안정됨.
    """
    gaps: list[PairingGap] = []
    campaigns = [c for c in campaigns if c.cost > 0 and c.bid_mode in ("auto", "manual")]

    def _key(c: CampaignPerf) -> tuple[str, str]:
        return (c.age_range, c.creative_id)

    manual = {_key(c): c for c in campaigns if c.bid_mode == "manual"}
    auto = {_key(c): c for c in campaigns if c.bid_mode == "auto"}

    for key, c in manual.items():
        if key not in auto:
            gaps.append(PairingGap(campaign=c.name, missing_counterpart="auto"))
    for key, c in auto.items():
        if key not in manual:
            gaps.append(PairingGap(campaign=c.name, missing_counterpart="manual"))

    return gaps


# 페어 CPA 차이가 이 비율 미만이면 '비슷함(안정화 중)'으로 본다.
_PAIR_TIE_THRESHOLD = 0.10


@dataclass(frozen=True)
class PairComparison:
    """동일 조건(연령+소재)의 수동/자동 페어 head-to-head 비교."""

    creative_key: str
    age_range: str
    manual: CampaignPerf
    auto: CampaignPerf

    @property
    def winner(self) -> Literal["manual", "auto", "tie"]:
        m, a = self.manual, self.auto
        # 행동이 없는 쪽은 효율 판단 불가 → 행동 있는 쪽 우위.
        if m.actions == 0 and a.actions == 0:
            return "tie"
        if m.actions == 0:
            return "auto"
        if a.actions == 0:
            return "manual"
        lo = min(m.cpa, a.cpa)
        if lo <= 0 or abs(m.cpa - a.cpa) / lo < _PAIR_TIE_THRESHOLD:
            return "tie"
        return "manual" if m.cpa < a.cpa else "auto"

    @property
    def cpa_gap_pct(self) -> float:
        """두 모드 CPA 차이 비율(%). 한쪽이라도 CPA가 없으면 0."""
        m, a = self.manual.cpa, self.auto.cpa
        lo = min(m, a)
        if m <= 0 or a <= 0 or lo <= 0:
            return 0.0
        return abs(m - a) / lo * 100

    @property
    def recommendation(self) -> str:
        gap = self.cpa_gap_pct
        if self.winner == "manual":
            return (
                f"수동이 CPA {gap:.0f}% 더 효율적이에요. 수동이 안정화된 단계라면 "
                "자동을 종료하고 수동에 집중하는 걸 검토해보세요."
            )
        if self.winner == "auto":
            return (
                f"자동이 CPA {gap:.0f}% 더 효율적이에요. 수동 비중을 줄이고 자동 위주로 "
                "확장하는 걸 검토해보세요."
            )
        return (
            "두 모드 효율이 비슷해요. 아직 안정화 중이니 페어를 유지하며 조금 더 "
            "지켜보세요. (변수는 입찰 모드 하나만 다르게 통제)"
        )


def compare_auto_manual_pairs(
    campaigns: Sequence[CampaignPerf],
) -> list[PairComparison]:
    """수동·자동이 모두 존재하는 페어(같은 연령+소재)를 head-to-head로 비교.

    누락 페어는 제외(그건 check_auto_manual_pairing이 담당). 비용이 있는 캠페인만 대상.
    CPA가 좋은(낮은) 쪽이 더 위에 오도록, 효율 격차가 큰 페어부터 정렬해 반환한다.
    """
    campaigns = [c for c in campaigns if c.cost > 0 and c.bid_mode in ("auto", "manual")]

    def _key(c: CampaignPerf) -> tuple[str, str]:
        return (c.age_range, c.creative_id)

    manual = {_key(c): c for c in campaigns if c.bid_mode == "manual"}
    auto = {_key(c): c for c in campaigns if c.bid_mode == "auto"}

    pairs = [
        PairComparison(
            creative_key=key[1] or m.name,
            age_range=key[0],
            manual=m,
            auto=auto[key],
        )
        for key, m in manual.items()
        if key in auto
    ]
    # 효율 격차 큰 페어 우선 (의사결정 임팩트 순).
    pairs.sort(key=lambda p: p.cpa_gap_pct, reverse=True)
    return pairs


# ───────────────────────── xlsx parser ─────────────────────────


def parse_demographic_xlsx(content: bytes) -> dict[str, list[Segment] | list[CampaignPerf]]:
    """Parse a 당근 광고 관리자 내보내기 xlsx.

    Supports two layouts:

    1. **Pre-aggregated per-sheet** (legacy/template format):
       - Sheet with '성별' column → gender segments
       - Sheet with '연령' or '연령대' column → age segments
       - Sheet with '캠페인' column → campaigns

    2. **Long-format breakdown** (당근 광고관리자 직접 내보내기):
       - Single sheet with rows = (기간 × 캠페인 × 연령) cross-tab.
       - Aggregated into age segments + campaign rows.
       - Bid mode parsed from campaign-name suffix (`_수동` / `_자동`).

    Returns a dict with keys "genders", "ages", "campaigns".
    Missing dimensions return empty lists.
    """
    import io

    import openpyxl  # local import so tests can stub if needed

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    result: dict[str, list] = {"genders": [], "ages": [], "campaigns": []}
    period_first: str | None = None
    period_last: str | None = None
    age_gender_accum: dict[tuple[str, str], dict[str, int]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = _find_header_row(rows)
        if header_idx < 0:
            continue
        header = [str(c or "").strip() for c in rows[header_idx]]

        # Try breakdown format first (single-sheet long-format)
        if _is_breakdown_format(header):
            data_rows = rows[header_idx + 1 :]
            ages, campaigns, p_first, p_last = _aggregate_breakdown(header, data_rows)
            result["ages"].extend(ages)
            result["campaigns"].extend(campaigns)
            # 성별 컬럼이 있으면 (연령 × 성별) 조인 셀도 누적 (히트맵용).
            _accumulate_age_gender(header, data_rows, age_gender_accum)
            if p_first and (period_first is None or p_first < period_first):
                period_first = p_first
            if p_last and (period_last is None or p_last > period_last):
                period_last = p_last
            continue

        kind = _detect_sheet_kind(header)
        if not kind:
            continue

        col_map = _map_demographic_columns(header, kind)
        if col_map.get("label") is None:
            continue

        for row in rows[header_idx + 1 :]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            label = _cell_str(row, col_map["label"])
            if not label or label in {"합계", "Total", "소계"}:
                continue

            cost = _cell_int(row, col_map.get("cost"))
            actions = _cell_int(row, col_map.get("actions"))
            impressions = _cell_int(row, col_map.get("impressions"))
            clicks = _cell_int(row, col_map.get("clicks"))

            if kind == "campaign":
                creatives = _cell_int(row, col_map.get("creative_count")) or 1
                bid_mode = _cell_str(row, col_map.get("bid_mode")).lower()
                if "자동" in bid_mode or "auto" in bid_mode:
                    bm = "auto"
                elif "수동" in bid_mode or "manual" in bid_mode:
                    bm = "manual"
                else:
                    bm = "unknown"
                result["campaigns"].append(
                    CampaignPerf(
                        name=label,
                        cost=cost,
                        actions=actions,
                        creative_count=creatives,
                        impressions=impressions,
                        clicks=clicks,
                        bid_mode=bm,
                        age_range=_cell_str(row, col_map.get("age_range")),
                        creative_id=_cell_str(row, col_map.get("creative_id")),
                    )
                )
            else:
                seg = Segment(
                    label=label,
                    cost=cost,
                    actions=actions,
                    impressions=impressions,
                    clicks=clicks,
                )
                result["genders" if kind == "gender" else "ages"].append(seg)

    result["age_gender_cells"] = _finalize_age_gender(age_gender_accum)
    result["meta"] = {
        "period_first": period_first or "",
        "period_last": period_last or "",
        "campaign_names": [c.name for c in result["campaigns"]],
    }
    return result


def _find_header_row(rows: list[tuple]) -> int:
    keywords = ("성별", "연령", "캠페인", "비용", "노출", "클릭", "행동")
    for idx, row in enumerate(rows[:15]):
        if not row:
            continue
        text = " ".join(str(c or "") for c in row)
        if sum(1 for kw in keywords if kw in text) >= 2:
            return idx
    return -1


def _detect_sheet_kind(header: list[str]) -> str | None:
    joined = " ".join(header)
    if "성별" in joined:
        return "gender"
    if "연령" in joined:
        return "age"
    if "캠페인" in joined:
        return "campaign"
    return None


def _map_demographic_columns(header: list[str], kind: str) -> dict[str, int | None]:
    m: dict[str, int | None] = {
        "label": None,
        "cost": None,
        "actions": None,
        "impressions": None,
        "clicks": None,
        "creative_count": None,
        "bid_mode": None,
        "age_range": None,
        "creative_id": None,
    }
    for idx, h in enumerate(header):
        normalized = h.replace(" ", "").replace("(원)", "").replace("(%)", "").replace("(회)", "")
        if m["label"] is None:
            if kind == "gender" and "성별" in normalized:
                m["label"] = idx
                continue
            if kind == "age" and ("연령" in normalized):
                m["label"] = idx
                continue
            if kind == "campaign" and ("캠페인" in normalized):
                m["label"] = idx
                continue
        if m["cost"] is None and ("비용" in normalized or "광고비" in normalized or "집행" in normalized):
            m["cost"] = idx
            continue
        if m["actions"] is None and ("총행동" in normalized or "행동수" in normalized or "문의" in normalized or "전환" in normalized):
            m["actions"] = idx
            continue
        if m["impressions"] is None and ("노출" in normalized):
            m["impressions"] = idx
            continue
        if m["clicks"] is None and ("클릭" in normalized):
            m["clicks"] = idx
            continue
        if m["creative_count"] is None and ("소재수" in normalized or "소재" == normalized):
            m["creative_count"] = idx
            continue
        if m["bid_mode"] is None and ("모드" in normalized or "자동수동" in normalized or "입찰" in normalized):
            m["bid_mode"] = idx
            continue
        if m["age_range"] is None and kind == "campaign" and "연령" in normalized:
            m["age_range"] = idx
            continue
        if m["creative_id"] is None and ("소재id" in normalized.lower() or "소재명" in normalized):
            m["creative_id"] = idx
    return m


def _cell_int(row: tuple, idx: int | None) -> int:
    if idx is None or idx >= len(row):
        return 0
    v = row[idx]
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    try:
        s = str(v).replace(",", "").replace("원", "").replace("%", "").strip()
        if not s or s in {"-", "nan"}:
            return 0
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _cell_str(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


# ───────────────────── long-format breakdown parser ─────────────────────

# Action columns: any of these contribute to "총행동" when summed per row.
# Source: 당근 광고관리자 직접 내보내기 (기간×캠페인×연령 breakdown).
_ACTION_KEYWORDS = (
    "단골", "후기", "쿠폰", "관심", "댓글",
    "전화 문의", "채팅 문의", "포장 주문", "리드폼 잠재고객",
)


def _is_breakdown_format(header: list[str]) -> bool:
    """Detect 당근 광고관리자 직접 내보내기 (long-format) layout.

    Signature: header contains 캠페인 + 연령 + 비용 columns simultaneously, which
    only happens in the breakdown sheet (cross-tabulated rows).
    """
    joined = " ".join(header)
    has_campaign = "캠페인" in joined
    has_age = "연령" in joined
    has_cost = "비용" in joined
    has_period = "기간" in joined or "날짜" in joined
    return has_campaign and has_age and has_cost and has_period


def _find_breakdown_col(header: list[str], *keywords: str) -> int | None:
    """Locate first column whose normalized header contains any keyword."""
    for idx, h in enumerate(header):
        normalized = h.replace(" ", "")
        for kw in keywords:
            if kw.replace(" ", "") in normalized:
                return idx
    return None


def _classify_bid_mode_from_name(name: str) -> Literal["auto", "manual", "unknown"]:
    """Parse 자동/수동 from campaign name suffix."""
    if not name:
        return "unknown"
    # match suffix or any token boundary
    if name.endswith("_자동") or "_자동_" in name or name.endswith(" 자동"):
        return "auto"
    if name.endswith("_수동") or "_수동_" in name or name.endswith(" 수동"):
        return "manual"
    if "자동" in name and "수동" not in name:
        return "auto"
    if "수동" in name and "자동" not in name:
        return "manual"
    return "unknown"


def _campaign_creative_key(name: str) -> str:
    """Extract the shared creative identity from a campaign name by
    stripping the trailing `_수동` / `_자동` token. Used for auto/manual pairing.
    """
    for suffix in ("_수동", "_자동", " 수동", " 자동"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name


def _aggregate_breakdown(
    header: list[str], data_rows: list[tuple]
) -> tuple[list[Segment], list[CampaignPerf], str | None, str | None]:
    """Aggregate long-format rows into age segments + per-campaign rows.

    Each input row is (기간, 캠페인, 캠페인ID?, 연령, 비용, 노출, 도달?, 클릭, ...,
    단골, 후기, 쿠폰, 관심, 댓글, 전화문의, 채팅문의, 포장주문, 리드폼 잠재고객, ...).

    Returns (ages, campaigns, period_first, period_last) — period strings are
    inclusive bounds extracted from the 기간/날짜 column (empty if not present).
    """
    col_age = _find_breakdown_col(header, "연령")
    col_campaign = _find_breakdown_col(header, "캠페인 이름", "캠페인이름", "캠페인")
    col_cost = _find_breakdown_col(header, "비용")
    col_imp = _find_breakdown_col(header, "노출")
    col_click = _find_breakdown_col(header, "클릭 수", "클릭수")
    col_period = _find_breakdown_col(header, "기간", "날짜")

    action_cols = [
        idx for idx, h in enumerate(header)
        if any(kw.replace(" ", "") in h.replace(" ", "") for kw in _ACTION_KEYWORDS)
        and "CPA" not in h and "CVR" not in h
        and "비용" not in h  # exclude '단골당 비용' etc.
    ]

    if col_age is None or col_cost is None:
        return [], [], None, None

    age_accum: dict[str, dict[str, int]] = {}
    camp_accum: dict[str, dict[str, int]] = {}
    period_first: str | None = None
    period_last: str | None = None

    for row in data_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        age = _cell_str(row, col_age)
        camp = _cell_str(row, col_campaign) if col_campaign is not None else ""

        if not age and not camp:
            continue
        if age in {"합계", "Total", "소계"}:
            continue

        if col_period is not None:
            period_val = _normalize_period_cell(row, col_period)
            if period_val:
                if period_first is None or period_val < period_first:
                    period_first = period_val
                if period_last is None or period_val > period_last:
                    period_last = period_val

        cost = _cell_int(row, col_cost)
        imp = _cell_int(row, col_imp) if col_imp is not None else 0
        clk = _cell_int(row, col_click) if col_click is not None else 0
        actions = sum(_cell_int(row, c) for c in action_cols)

        if age:
            acc = age_accum.setdefault(
                age, {"cost": 0, "actions": 0, "impressions": 0, "clicks": 0},
            )
            acc["cost"] += cost
            acc["actions"] += actions
            acc["impressions"] += imp
            acc["clicks"] += clk

        if camp:
            acc = camp_accum.setdefault(
                camp,
                {"cost": 0, "actions": 0, "impressions": 0, "clicks": 0,
                 "ages": set()},  # type: ignore[dict-item]
            )
            acc["cost"] += cost
            acc["actions"] += actions
            acc["impressions"] += imp
            acc["clicks"] += clk
            if age:
                acc["ages"].add(age)  # type: ignore[union-attr]

    ages = [
        Segment(
            label=label, cost=v["cost"], actions=v["actions"],
            impressions=v["impressions"], clicks=v["clicks"],
        )
        for label, v in sorted(age_accum.items(), key=lambda kv: _age_sort_key(kv[0]))
    ]

    campaigns = []
    for name, v in camp_accum.items():
        ages_set: set[str] = v["ages"]  # type: ignore[assignment]
        age_range = ",".join(sorted(ages_set, key=_age_sort_key)) if ages_set else ""
        campaigns.append(
            CampaignPerf(
                name=name,
                cost=v["cost"], actions=v["actions"],
                impressions=v["impressions"], clicks=v["clicks"],
                creative_count=1,
                bid_mode=_classify_bid_mode_from_name(name),
                age_range=age_range,
                creative_id=_campaign_creative_key(name),
            )
        )

    return ages, campaigns, period_first, period_last


# ───────────────────── age × gender joint (heatmap) ─────────────────────


@dataclass(frozen=True)
class AgeGenderCell:
    """A joint (연령 × 성별) cell aggregated from breakdown rows."""

    age: str
    gender: str
    cost: int = 0
    actions: int = 0
    impressions: int = 0
    clicks: int = 0

    @property
    def cpa(self) -> float:
        return self.cost / self.actions if self.actions else 0.0

    @property
    def ctr(self) -> float:
        return (self.clicks / self.impressions * 100) if self.impressions else 0.0


def _gender_sort_key(label: str) -> tuple[int, str]:
    """남성 → 여성 → 기타 순으로 정렬."""
    order = {"남성": 0, "남": 0, "여성": 1, "여": 1}
    return (order.get(label.strip(), 2), label)


def _accumulate_age_gender(
    header: list[str], data_rows: list[tuple], accum: dict[tuple[str, str], dict[str, int]]
) -> None:
    """breakdown 행에서 (연령, 성별) 조인 셀을 accum에 누적(in-place).

    성별 컬럼이 없으면 아무것도 하지 않는다(조인 불가 → 히트맵 미생성).
    """
    col_age = _find_breakdown_col(header, "연령")
    col_gender = _find_breakdown_col(header, "성별")
    col_cost = _find_breakdown_col(header, "비용")
    col_imp = _find_breakdown_col(header, "노출")
    col_click = _find_breakdown_col(header, "클릭 수", "클릭수")
    action_cols = [
        idx for idx, h in enumerate(header)
        if any(kw.replace(" ", "") in h.replace(" ", "") for kw in _ACTION_KEYWORDS)
        and "CPA" not in h and "CVR" not in h and "비용" not in h
    ]
    if col_age is None or col_gender is None or col_cost is None:
        return

    skip = {"합계", "Total", "소계", "전체", "알 수 없음"}
    for row in data_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        age = _cell_str(row, col_age)
        gender = _cell_str(row, col_gender)
        if not age or not gender or age in skip or gender in skip:
            continue
        acc = accum.setdefault(
            (age, gender), {"cost": 0, "actions": 0, "impressions": 0, "clicks": 0}
        )
        acc["cost"] += _cell_int(row, col_cost)
        acc["impressions"] += _cell_int(row, col_imp) if col_imp is not None else 0
        acc["clicks"] += _cell_int(row, col_click) if col_click is not None else 0
        acc["actions"] += sum(_cell_int(row, c) for c in action_cols)


def _finalize_age_gender(accum: dict[tuple[str, str], dict[str, int]]) -> dict:
    """누적 dict을 {cells, ages, genders} 렌더 구조로 마감."""
    cells = [
        AgeGenderCell(age=age, gender=gender, **vals)
        for (age, gender), vals in accum.items()
    ]
    ages = sorted({c.age for c in cells}, key=_age_sort_key)
    genders = sorted({c.gender for c in cells}, key=_gender_sort_key)
    return {"cells": cells, "ages": ages, "genders": genders}


def aggregate_age_gender_cells(header: list[str], data_rows: list[tuple]) -> dict:
    """breakdown 시트에서 (연령 × 성별) 조인 셀을 집계해 {cells, ages, genders} 반환.

    성별 컬럼이 없으면 빈 결과({cells: [], ...})를 반환한다.
    """
    accum: dict[tuple[str, str], dict[str, int]] = {}
    _accumulate_age_gender(header, data_rows, accum)
    return _finalize_age_gender(accum)


def _normalize_period_cell(row: tuple, idx: int | None) -> str:
    """Format a period cell value as YYYY.MM.DD-like string for sortable comparison."""
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    # datetime → ISO date
    try:
        from datetime import datetime, date
        if isinstance(v, datetime):
            return v.strftime("%Y.%m.%d")
        if isinstance(v, date):
            return v.strftime("%Y.%m.%d")
    except Exception:  # noqa: BLE001
        pass
    s = str(v).strip().rstrip(".")
    return s


# ─────────────────────────── funnel + economics ───────────────────────────


@dataclass(frozen=True)
class Funnel:
    """Aggregated funnel for ad → click → action with bottleneck identification."""

    impressions: int
    clicks: int
    actions: int

    @property
    def ctr(self) -> float:
        return (self.clicks / self.impressions * 100) if self.impressions > 0 else 0.0

    @property
    def cvr(self) -> float:
        return (self.clicks > 0 and (self.actions / self.clicks * 100)) or 0.0

    @property
    def drop_impression_to_click(self) -> float:
        if self.impressions <= 0:
            return 0.0
        return (1 - (self.clicks / self.impressions)) * 100

    @property
    def drop_click_to_action(self) -> float:
        if self.clicks <= 0:
            return 0.0
        return (1 - (self.actions / self.clicks)) * 100

    @property
    def bottleneck(self) -> str:
        """Returns the stage with highest drop-off, or '' if no data."""
        if self.impressions <= 0:
            return ""
        a = self.drop_impression_to_click
        b = self.drop_click_to_action
        if a >= b:
            return "노출→클릭"
        return "클릭→행동"


def calc_funnel(items: Sequence) -> Funnel:
    """Aggregate impressions/clicks/actions across Segments or CampaignPerf rows."""
    imps = sum(getattr(x, "impressions", 0) for x in items)
    clks = sum(getattr(x, "clicks", 0) for x in items)
    acts = sum(getattr(x, "actions", 0) for x in items)
    return Funnel(impressions=imps, clicks=clks, actions=acts)


@dataclass(frozen=True)
class Economics:
    """MAX CPA / 한계 소진율 / 손익 판정.

    avg_order_value · target_margin_rate · variable_cost_rate가 0이면
    아무 계산도 못 하므로 max_cpa = 0, status = "unknown"으로 반환된다.
    """

    total_cost: int
    total_actions: int
    current_cpa: float
    avg_order_value: int            # 객단가 (광고주 입력)
    target_margin_rate: float       # 0.0 ~ 1.0 (입력 % / 100)
    variable_cost_rate: float = 0.0  # 원가율 (0 = 디지털 상품 등)
    breakeven_cpa: float = 0.0       # 손익분기 CPA = 객단가 × (1 - 원가율)
    max_cpa: float = 0.0             # 목표이익 반영 후 허용 CPA
    burn_rate: float = 0.0           # current_cpa / max_cpa
    status: str = "unknown"          # profit | breakeven | loss | unknown
    expected_revenue: int = 0
    expected_profit: int = 0


def calc_economics(
    total_cost: int,
    total_actions: int,
    *,
    avg_order_value: int = 0,
    target_margin_rate: float = 0.0,
    variable_cost_rate: float = 0.0,
) -> Economics:
    """Calculate MAX CPA + 손익 status from inputs.

    breakeven_cpa = 객단가 × (1 - 원가율)
    max_cpa       = breakeven_cpa × (1 - 목표이익률)
    burn_rate     = current_cpa / max_cpa
        status:
          burn_rate ≥ 1.0 → "loss"  (적자)
          burn_rate ≥ 0.7 → "breakeven"  (손익분기 근접)
          burn_rate <  0.7 → "profit"  (확장 여력)
    """
    current_cpa = total_cost / total_actions if total_actions > 0 else 0.0
    if avg_order_value <= 0:
        return Economics(
            total_cost=total_cost, total_actions=total_actions,
            current_cpa=current_cpa, avg_order_value=0,
            target_margin_rate=target_margin_rate,
            variable_cost_rate=variable_cost_rate,
            status="unknown",
        )

    breakeven = avg_order_value * (1.0 - max(0.0, min(1.0, variable_cost_rate)))
    margin = max(0.0, min(1.0, target_margin_rate))
    max_cpa = breakeven * (1.0 - margin)

    burn = current_cpa / max_cpa if max_cpa > 0 else 0.0
    if burn >= 1.0:
        status = "loss"
    elif burn >= 0.7:
        status = "breakeven"
    else:
        status = "profit"

    revenue = total_actions * avg_order_value
    profit = int(revenue * (1.0 - variable_cost_rate) - total_cost)

    return Economics(
        total_cost=total_cost,
        total_actions=total_actions,
        current_cpa=current_cpa,
        avg_order_value=avg_order_value,
        target_margin_rate=margin,
        variable_cost_rate=variable_cost_rate,
        breakeven_cpa=breakeven,
        max_cpa=max_cpa,
        burn_rate=burn,
        status=status,
        expected_revenue=revenue,
        expected_profit=profit,
    )


def _age_sort_key(label: str) -> tuple[int, str]:
    """Sort age labels like '40-44', '45-49', '60 이상' numerically."""
    import re
    m = re.match(r"\s*(\d+)", label)
    if m:
        return (int(m.group(1)), label)
    return (10_000, label)
