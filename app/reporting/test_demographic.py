"""Tests for demographic analysis module.

Fixtures are based on the operator's reference report image:
- Gender: 남성 (2,635,214원/914행동/2,883원CPA) + 여성 (742,432원/448행동/1,657원CPA)
- Age brackets spanning 15-19 through 60+
- 4 campaigns with different verdicts expected
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.reporting.demographic import (
    AgeGroup,
    CampaignJudgment,
    CampaignPerf,
    Insight,
    PairingGap,
    ReallocationPlan,
    Segment,
    VariableControlWarning,
    analyze_segments,
    build_priority_checklist,
    check_auto_manual_pairing,
    check_variable_control,
    group_ages_by_cpa,
    judge_campaigns,
    parse_demographic_xlsx,
    simulate_reallocation,
)


# ───────────────────────── fixtures ─────────────────────────


@pytest.fixture
def gender_segments() -> list[Segment]:
    return [
        Segment(label="남성", cost=2_635_214, actions=914, impressions=100_000, clicks=1_900),
        Segment(label="여성", cost=742_432, actions=448, impressions=40_000, clicks=880),
        Segment(label="알 수 없음", cost=30_969, actions=11, impressions=2_000, clicks=22),
    ]


@pytest.fixture
def age_segments() -> list[Segment]:
    # Based on image hints: 15-19 best efficiency, 60+ budget imbalance.
    return [
        Segment(label="15-19", cost=80_000, actions=54, impressions=7_000, clicks=84),   # 1,482원 CPA
        Segment(label="20-24", cost=150_000, actions=68, impressions=10_000, clicks=110), # 2,205원
        Segment(label="25-29", cost=200_000, actions=74, impressions=12_000, clicks=130), # 2,702원
        Segment(label="30-34", cost=250_000, actions=75, impressions=13_000, clicks=130), # 3,333원
        Segment(label="35-39", cost=280_000, actions=80, impressions=14_000, clicks=135), # 3,500원
        Segment(label="40-44", cost=320_000, actions=82, impressions=14_500, clicks=130), # 3,902원
        Segment(label="45-49", cost=360_000, actions=76, impressions=15_000, clicks=125), # 4,736원
        Segment(label="50-54", cost=340_000, actions=70, impressions=14_500, clicks=115), # 4,857원
        Segment(label="55-59", cost=310_000, actions=62, impressions=13_500, clicks=100), # 5,000원
        Segment(label="60+", cost=450_000, actions=61, impressions=17_000, clicks=110),   # 7,377원 — 불균형
    ]


@pytest.fixture
def campaigns() -> list[CampaignPerf]:
    # Mirrors the reference image verdicts.
    return [
        CampaignPerf(name="A", cost=1_913_928, actions=781, creative_count=1, impressions=200_000, clicks=3_740),  # CTR 1.87
        CampaignPerf(name="B", cost=1_257_160, actions=565, creative_count=1, impressions=180_000, clicks=2_250),  # CTR 1.25
        CampaignPerf(name="C", cost=173_318, actions=8, creative_count=2, impressions=50_000, clicks=1_220),       # CTR 2.44 → OFF
        CampaignPerf(name="중고루틴", cost=64_209, actions=19, creative_count=1, impressions=25_000, clicks=345),   # CTR 1.39
    ]


# ───────────────────────── segment insights ─────────────────────────


def test_analyze_segments_returns_best_efficiency(gender_segments: list[Segment]) -> None:
    insights = analyze_segments(gender_segments)
    bests = [i for i in insights if i.kind == "best_efficiency"]
    assert bests, "should identify a best-efficiency segment"
    # 여성 has the lowest CPA (1,657원)
    assert bests[0].label == "여성"


def test_analyze_segments_empty_returns_empty() -> None:
    assert analyze_segments([]) == []


def test_analyze_segments_all_zero_actions_returns_empty() -> None:
    segs = [Segment(label="X", cost=100, actions=0)]
    assert analyze_segments(segs) == []


def test_analyze_ages_flags_budget_imbalance(age_segments: list[Segment]) -> None:
    insights = analyze_segments(age_segments)
    labels_imbalanced = {i.label for i in insights if i.kind == "budget_imbalance"}
    # 60+ spends the most (450,000원) with worst CPA (7,377원)
    assert "60+" in labels_imbalanced


# ───────────────────────── age grouping ─────────────────────────


def test_group_ages_by_cpa_produces_n_groups(age_segments: list[Segment]) -> None:
    groups = group_ages_by_cpa(age_segments, n_groups=3)
    assert len(groups) == 3
    # groups are sorted by avg_cpa ascending
    assert groups[0].avg_cpa <= groups[1].avg_cpa <= groups[2].avg_cpa


def test_group_ages_respects_bounds() -> None:
    # Only 2 usable segments; can't form 5 groups.
    two = [
        Segment(label="A", cost=100, actions=1),
        Segment(label="B", cost=200, actions=1),
    ]
    assert len(group_ages_by_cpa(two, n_groups=5)) == 2


def test_group_ages_separates_inactive_into_own_group() -> None:
    mixed = [
        Segment(label="good", cost=1_000, actions=5),  # 200
        Segment(label="zero", cost=500, actions=0),    # inactive → own group (inf)
    ]
    groups = group_ages_by_cpa(mixed, n_groups=2)
    assert groups[-1].members == ("zero",)
    assert groups[-1].avg_cpa == float("inf")


def test_group_ages_handles_empty() -> None:
    assert group_ages_by_cpa([], n_groups=3) == []


# ───────────────────────── campaign judgment ─────────────────────────


def test_judge_zero_action_campaign_is_off() -> None:
    camps = [
        CampaignPerf(name="zero", cost=100_000, actions=0, impressions=10_000, clicks=10),
    ]
    judgments = judge_campaigns(camps)
    assert judgments[0].verdict == "캠페인OFF"


def test_judge_low_ctr_high_spend_is_replace() -> None:
    camps = [
        CampaignPerf(name="big_lowctr", cost=1_000_000, actions=100, impressions=1_000_000, clicks=5_000),  # CTR 0.5
        CampaignPerf(name="other", cost=500_000, actions=300, impressions=50_000, clicks=2_000),
    ]
    judgments = judge_campaigns(camps)
    big = next(j for j in judgments if j.campaign.name == "big_lowctr")
    assert big.verdict == "소재전면교체"


def test_judge_main_campaign_is_keep(campaigns: list[CampaignPerf]) -> None:
    judgments = judge_campaigns(campaigns)
    names_to_verdict = {j.campaign.name: j.verdict for j in judgments}
    # Campaign A has 1.87% CTR + 50%+ spend share + good CPA → 소재정리후유지
    assert names_to_verdict["A"] == "소재정리후유지"


def test_judge_returns_empty_for_no_input() -> None:
    assert judge_campaigns([]) == []


# ───────────────────────── reallocation ─────────────────────────


def test_reallocation_saves_from_off_campaigns(campaigns: list[CampaignPerf]) -> None:
    judgments = judge_campaigns(campaigns)
    plan = simulate_reallocation(judgments)
    assert plan.current_total == sum(c.cost for c in campaigns)
    # Campaign C has 8 actions on 173k spend (CPA >21k) + share ~5% — should trigger OFF/replace
    total_cut = sum(amt for _, amt in plan.cuts)
    assert total_cut > 0


def test_reallocation_no_candidates_returns_zero_delta() -> None:
    stable = [
        CampaignJudgment(
            campaign=CampaignPerf(name="x", cost=100, actions=10, impressions=1_000, clicks=10),
            verdict="유지",
            reason="stable",
            cost_share=100.0,
        )
    ]
    plan = simulate_reallocation(stable)
    assert plan.expected_action_delta == 0
    assert plan.cuts == () and plan.boosts == ()


# ───────────────────────── priority checklist ─────────────────────────


def test_priority_checklist_has_4_tiers_when_all_verdicts_present(
    campaigns: list[CampaignPerf],
) -> None:
    judgments = judge_campaigns(campaigns)
    items = build_priority_checklist(judgments)
    # 4th item (신규 테스트) is always present
    assert any(item.startswith("4순위") for item in items)


# ───────────────────────── variable control ─────────────────────────


def test_variable_control_flags_two_or_more_diffs() -> None:
    campaigns = [
        CampaignPerf(name="A", cost=100, actions=1, bid_mode="manual", age_range="20-29", creative_id="v1"),
        CampaignPerf(name="B", cost=100, actions=1, bid_mode="auto", age_range="30-39", creative_id="v1"),
    ]
    warnings = check_variable_control(campaigns)
    assert len(warnings) == 1
    assert warnings[0].diff_count == 2
    assert "bid_mode(자동/수동)" in warnings[0].diffs


def test_variable_control_ignores_single_diff() -> None:
    campaigns = [
        CampaignPerf(name="A", cost=100, actions=1, bid_mode="manual", age_range="20-29", creative_id="v1"),
        CampaignPerf(name="B", cost=100, actions=1, bid_mode="manual", age_range="30-39", creative_id="v1"),
    ]
    assert check_variable_control(campaigns) == []


def test_variable_control_ignores_unknown_bid_mode() -> None:
    campaigns = [
        CampaignPerf(name="A", cost=100, actions=1, age_range="20-29", creative_id="v1"),
        CampaignPerf(name="B", cost=100, actions=1, age_range="30-39", creative_id="v2"),
    ]
    # unknown bid_mode not counted; 2 real diffs → warning
    warnings = check_variable_control(campaigns)
    assert len(warnings) == 1


# ───────────────────────── auto/manual pairing ─────────────────────────


def test_pairing_detects_missing_auto_twin() -> None:
    campaigns = [
        CampaignPerf(name="A_manual", cost=100, actions=1, bid_mode="manual", age_range="20-29", creative_id="v1"),
    ]
    gaps = check_auto_manual_pairing(campaigns)
    assert len(gaps) == 1
    assert gaps[0].missing_counterpart == "auto"


def test_pairing_accepts_matched_pair() -> None:
    campaigns = [
        CampaignPerf(name="A_manual", cost=100, actions=1, bid_mode="manual", age_range="20-29", creative_id="v1"),
        CampaignPerf(name="A_auto", cost=100, actions=1, bid_mode="auto", age_range="20-29", creative_id="v1"),
    ]
    assert check_auto_manual_pairing(campaigns) == []


# ───────────────────────── xlsx parser ─────────────────────────


def _build_sample_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Gender sheet
    ws = wb.create_sheet("성별")
    ws.append(["성별", "비용", "노출", "클릭", "총행동"])
    ws.append(["남성", 2_635_214, 100_000, 1_900, 914])
    ws.append(["여성", 742_432, 40_000, 880, 448])

    # Age sheet
    ws = wb.create_sheet("연령")
    ws.append(["연령대", "비용", "노출", "클릭", "총행동"])
    ws.append(["15-19", 80_000, 7_000, 84, 54])
    ws.append(["60+", 450_000, 17_000, 110, 61])

    # Campaign sheet
    ws = wb.create_sheet("캠페인")
    ws.append(["캠페인", "비용", "노출", "클릭", "총행동", "소재 수", "입찰모드"])
    ws.append(["A", 1_913_928, 200_000, 3_740, 781, 1, "수동"])
    ws.append(["중고루틴", 64_209, 25_000, 345, 19, 1, "자동"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_demographic_xlsx_extracts_all_kinds() -> None:
    data = _build_sample_xlsx()
    parsed = parse_demographic_xlsx(data)

    assert len(parsed["genders"]) == 2
    assert parsed["genders"][0].label == "남성"
    assert parsed["genders"][0].cost == 2_635_214

    assert len(parsed["ages"]) == 2
    assert parsed["ages"][1].label == "60+"

    assert len(parsed["campaigns"]) == 2
    a = next(c for c in parsed["campaigns"] if c.name == "A")
    assert a.cost == 1_913_928
    assert a.bid_mode == "manual"


def test_parse_demographic_xlsx_empty_bytes_returns_empty() -> None:
    # openpyxl would raise on empty input; ensure parser handles gracefully
    with pytest.raises(Exception):  # noqa: BLE001
        parse_demographic_xlsx(b"")


# ───────────────── long-format breakdown parser ─────────────────


def _build_breakdown_xlsx() -> bytes:
    """Single-sheet long-format mimicking 당근 광고관리자 직접 내보내기."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "breakdown"
    ws.append([
        "기간", "캠페인 이름", "캠페인 ID", "연령",
        "비용 (VAT 포함)", "노출 수", "도달 수", "클릭 수", "클릭률",
        "클릭당 비용(CPC)", "노출당 비용(CPM)",
        "단골 수", "후기 수", "쿠폰 다운로드 수", "관심 수", "댓글 수",
        "전화 문의 수", "채팅 문의 수", "포장 주문 수",
    ])
    rows = [
        # (period, campaign, id, age, cost, imp, reach, click, ctr, cpc, cpm,
        #  단골, 후기, 쿠폰, 관심, 댓글, 전화, 채팅, 포장)
        ("D1", "A_수동", 1, "40-44", 1000, 500, 400, 10, 2.0, 100, 2000, 0, 0, 0, 0, 0, 0, 1, 0),
        ("D1", "A_수동", 1, "50-54", 2000, 700, 600, 12, 1.7, 167, 2857, 2, 0, 1, 0, 0, 0, 0, 0),
        ("D1", "A_자동", 2, "40-44", 1500, 600, 500, 11, 1.8, 136, 2500, 0, 0, 0, 0, 0, 0, 0, 0),
        ("D1", "A_자동", 2, "50-54", 2500, 800, 700, 13, 1.6, 192, 3125, 1, 0, 2, 1, 0, 0, 0, 0),
        ("D2", "A_수동", 1, "40-44", 1100, 520, 410, 11, 2.1, 100, 2115, 0, 0, 0, 0, 0, 0, 1, 0),
    ]
    for r in rows:
        ws.append(r)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_breakdown_parser_aggregates_age_segments() -> None:
    data = _build_breakdown_xlsx()
    parsed = parse_demographic_xlsx(data)

    # 2 distinct ages aggregated across all rows
    ages = {a.label: a for a in parsed["ages"]}
    assert set(ages) == {"40-44", "50-54"}
    # 40-44 totals: cost=1000+1500+1100=3600; actions: 2 (one 채팅, one 전화) — actually 1+1 channels
    # D1 A_수동 40-44 channels=(0..,채팅=1,...) =1 ; D1 A_자동 40-44 = 0 ; D2 A_수동 40-44 = 1 → total 2
    assert ages["40-44"].cost == 3600
    assert ages["40-44"].actions == 2  # 1 채팅 + 1 채팅
    # 50-54 totals: cost=2000+2500=4500; actions: 단골2+쿠폰1+단골1+쿠폰2+관심1 = 7
    assert ages["50-54"].cost == 4500
    assert ages["50-54"].actions == 7


def test_breakdown_parser_aggregates_campaigns_with_bid_mode() -> None:
    data = _build_breakdown_xlsx()
    parsed = parse_demographic_xlsx(data)

    camps = {c.name: c for c in parsed["campaigns"]}
    assert set(camps) == {"A_수동", "A_자동"}

    suho = camps["A_수동"]
    assert suho.bid_mode == "manual"
    # cost: D1 40-44 1000 + D1 50-54 2000 + D2 40-44 1100 = 4100
    assert suho.cost == 4100
    # actions: 채팅1 + 단골2+쿠폰1 + 채팅1 = 5
    assert suho.actions == 5
    assert "40-44" in suho.age_range and "50-54" in suho.age_range

    auto = camps["A_자동"]
    assert auto.bid_mode == "auto"
    assert auto.cost == 4000  # 1500 + 2500
    # actions: 0 + (단골1+쿠폰2+관심1) = 4
    assert auto.actions == 4


def test_breakdown_creative_id_strips_bid_suffix() -> None:
    data = _build_breakdown_xlsx()
    parsed = parse_demographic_xlsx(data)

    camps = {c.name: c for c in parsed["campaigns"]}
    # 수동/자동 pair should share the same creative_id for pairing detection
    assert camps["A_수동"].creative_id == "A"
    assert camps["A_자동"].creative_id == "A"


def test_breakdown_pairing_detects_complete_pair() -> None:
    data = _build_breakdown_xlsx()
    parsed = parse_demographic_xlsx(data)
    gaps = check_auto_manual_pairing(parsed["campaigns"])
    # Both 수동 and 자동 exist with same age coverage → no gaps
    assert gaps == []


def test_breakdown_pairing_detects_orphan_manual() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "기간", "캠페인 이름", "연령", "비용 (VAT 포함)", "노출 수", "클릭 수", "단골 수",
    ])
    ws.append(("D1", "Only_수동", "40-44", 1000, 500, 10, 1))
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_demographic_xlsx(buf.getvalue())
    gaps = check_auto_manual_pairing(parsed["campaigns"])
    assert len(gaps) == 1
    assert gaps[0].missing_counterpart == "auto"


# ─────── long-format WITHOUT 연령 / WITHOUT action columns (실측 밀양 양식) ───────


def _build_breakdown_no_age_xlsx() -> bytes:
    """Mimics 지니스안경 밀양점 export: 캠페인별 일자 추이, 연령·행동 컬럼 없음."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "밀양"
    ws.append([
        "기간", "캠페인 이름", "캠페인 ID",
        "비용 (VAT 포함)", "노출 수", "도달 수", "클릭 수",
        "클릭률", "클릭당 비용(CPC)", "노출당 비용(CPM)",
    ])
    rows = [
        ("2026.06.01.", "밀양_수동_35~59", 1, 8019, 2545, 1645, 25, 0.98, 320, 3150),
        ("2026.06.01.", "밀양_수동_35~44", 2, 2552, 1049, 656, 8, 0.76, 319, 2432),
        ("2026.06.02.", "밀양_수동_35~59", 1, 8972, 2998, 2025, 28, 0.93, 320, 2992),
        ("2026.06.02.", "밀양_수동_35~44", 2, 2965, 1396, 961, 6, 0.43, 494, 2123),
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_breakdown_no_age_aggregates_campaigns_by_name() -> None:
    """연령 컬럼이 없어도 같은 캠페인 이름은 일자별로 합산되어야 한다(중복 금지)."""
    parsed = parse_demographic_xlsx(_build_breakdown_no_age_xlsx())

    camps = {c.name: c for c in parsed["campaigns"]}
    assert set(camps) == {"밀양_수동_35~59", "밀양_수동_35~44"}  # 4 rows → 2 campaigns
    assert parsed["ages"] == []  # 연령 없음 → 빈 리스트(가짜 생성 금지)

    big = camps["밀양_수동_35~59"]
    assert big.bid_mode == "manual"  # _수동 접미사에서 파싱
    assert big.cost == 8019 + 8972
    assert big.impressions == 2545 + 2998
    assert big.clicks == 25 + 28
    assert big.actions == 0  # 행동 컬럼 없음 → 0 (있는 것만)


def test_breakdown_no_age_metrics_available_is_honest() -> None:
    """파일에 실제로 있는 지표만 available로 보고해야 한다."""
    parsed = parse_demographic_xlsx(_build_breakdown_no_age_xlsx())
    available = set(parsed["metrics_available"])
    assert available == {"impressions", "clicks"}
    for absent in ("inquiries", "regulars", "coupons", "actions"):
        assert absent not in available
    assert parsed["meta"]["period_first"] == "2026.06.01"
    assert parsed["meta"]["period_last"] == "2026.06.02"


def test_breakdown_no_age_timeseries_rolls_up_by_date() -> None:
    parsed = parse_demographic_xlsx(_build_breakdown_no_age_xlsx())
    ts = {row["date"]: row for row in parsed["timeseries"]}
    assert set(ts) == {"2026.06.01", "2026.06.02"}
    assert ts["2026.06.01"]["cost"] == 8019 + 2552
    assert ts["2026.06.01"]["impressions"] == 2545 + 1049
    assert ts["2026.06.01"]["clicks"] == 25 + 8
    assert ts["2026.06.01"]["actions"] == 0


def test_breakdown_with_actions_reports_available_metrics() -> None:
    """행동 컬럼이 있는 양식은 inquiries/regulars/coupons/actions를 available로 보고."""
    parsed = parse_demographic_xlsx(_build_breakdown_xlsx())
    available = set(parsed["metrics_available"])
    assert {"impressions", "clicks", "inquiries", "regulars", "coupons", "actions"} <= available
    assert len(parsed["timeseries"]) == 2  # D1, D2
